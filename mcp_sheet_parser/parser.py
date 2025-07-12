# parser.py
# 表格解析核心逻辑

import os
import pandas as pd
import openpyxl
import xlrd
from pandas.errors import EmptyDataError
from openpyxl.styles.colors import Color
from .config import Config, ErrorMessages
from .utils import setup_logger, validate_file_path, get_file_extension, is_supported_format, clean_cell_value
from .security import validate_file_path as security_validate_file_path
from .performance import (
    StreamingExcelParser, PerformanceOptimizer, MemoryMonitor,
    create_progress_callback, benchmark_performance
)
from .formula_processor import FormulaProcessor
from .chart_converter import ChartConverter
from .exceptions import (
    ErrorHandler, ErrorContext, ErrorSeverity,
    FileNotFoundError, UnsupportedFormatError, FileSizeExceededError,
    WorksheetParsingError, CellProcessingError, ParsingError,
    error_handler, safe_execute, create_error_context
)


def _get_cell_style(cell):
    style = {}
    # 字体
    if cell.font:
        style['bold'] = cell.font.bold
        style['italic'] = cell.font.italic
        style['font_size'] = cell.font.sz
        style['font_name'] = cell.font.name
        style['underline'] = cell.font.underline
        style['strike'] = cell.font.strike
        # 字体颜色
        if cell.font.color and cell.font.color.type == 'rgb' and cell.font.color.rgb:
            style['font_color'] = f"#{cell.font.color.rgb[-6:]}"
    # 背景色
    if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != '00000000':
        style['bg_color'] = f"#{cell.fill.fgColor.rgb[-6:]}"
    # 对齐
    if cell.alignment:
        if cell.alignment.horizontal:
            style['align'] = cell.alignment.horizontal
        if cell.alignment.vertical:
            style['valign'] = cell.alignment.vertical
        style['wrap_text'] = cell.alignment.wrap_text
    # 边框
    if cell.border:
        border = {}
        for side in ['top', 'bottom', 'left', 'right']:
            border_side = getattr(cell.border, side)
            if border_side and border_side.style and border_side.style != 'none':
                # 安全处理边框颜色
                border_color = '#000000'  # 默认黑色
                if border_side.color and border_side.color.rgb:
                    try:
                        # 处理RGB对象
                        rgb_value = border_side.color.rgb
                        if hasattr(rgb_value, '__getitem__'):  # 可索引对象
                            border_color = f"#{rgb_value[-6:]}"
                        elif hasattr(rgb_value, 'value'):  # RGB对象可能有value属性
                            border_color = f"#{rgb_value.value[-6:]}"
                        else:
                            border_color = f"#{str(rgb_value)[-6:]}"
                    except (TypeError, AttributeError, IndexError):
                        border_color = '#000000'  # 发生错误时使用默认颜色
                
                border[side] = {
                    'style': border_side.style,
                    'color': border_color
                }
        if border:
            style['border'] = border
    # 超链接
    if cell.hyperlink:
        style['hyperlink'] = str(cell.hyperlink.target)
    # 批注/注释
    if cell.comment:
        style['comment'] = cell.comment.text
    return style


class SheetParser:
    def __init__(self, file_path, config=None, progress_callback=None):
        self.file_path = file_path
        self.config = config or Config()
        self.progress_callback = progress_callback
        self.logger = setup_logger(__name__)
        self.sheets = []  # 存储所有sheet的结构化数据
        self.error_handler = ErrorHandler(self.logger)
        
        # 初始化性能优化组件
        self.performance_optimizer = PerformanceOptimizer()
        self.memory_monitor = MemoryMonitor(
            max_memory_mb=getattr(self.config, 'MAX_MEMORY_MB', 2048)
        ) if getattr(self.config, 'ENABLE_PERFORMANCE_MODE', False) else None
        
        # 初始化公式处理器
        self.formula_processor = FormulaProcessor(config) if getattr(self.config, 'ENABLE_FORMULA_PROCESSING', True) else None
        
        # 初始化图表转换器
        self.chart_converter = ChartConverter(config, file_path) if getattr(self.config, 'ENABLE_CHART_CONVERSION', True) else None
        
        # 验证文件
        self._validate_file()
        
    def _validate_file(self):
        """验证文件的完整流程"""
        context = create_error_context("文件验证", file_path=self.file_path)
        
        # 检查文件是否存在
        if not validate_file_path(self.file_path):
            raise FileNotFoundError(self.file_path, context=context)
        
        # 安全检查
        is_safe, warnings = security_validate_file_path(self.file_path)
        if not is_safe:
            from .exceptions import SecurityError
            raise SecurityError(f"文件安全检查失败: {'; '.join(warnings)}", context=context)
        
        # 检查文件格式支持
        if not is_supported_format(self.file_path):
            ext = get_file_extension(self.file_path)
            raise UnsupportedFormatError(self.file_path, ext, context=context)
        
        # 检查文件大小
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        if file_size_mb > self.config.MAX_FILE_SIZE_MB:
            raise FileSizeExceededError(
                self.file_path, file_size_mb, self.config.MAX_FILE_SIZE_MB, context=context
            )
        
        # 为大文件优化性能配置
        if file_size_mb > 100 and getattr(self.config, 'ENABLE_PERFORMANCE_MODE', False):
            optimized_config = self.performance_optimizer.optimize_for_file(self.file_path)
            self.config.CHUNK_SIZE = getattr(optimized_config, 'CHUNK_SIZE', 1000)
            self.logger.info(f"启用性能优化模式，块大小: {self.config.CHUNK_SIZE}")
            
            if self.memory_monitor:
                self.logger.info(self.memory_monitor.get_memory_stats())

    def parse(self):
        """
        解析表格文件，返回结构化数据
        """
        context = create_error_context("文件解析", file_path=self.file_path)
        
        try:
            ext = get_file_extension(self.file_path)
            self.logger.info(f"开始解析文件: {self.file_path} (格式: {ext})")
            
            if ext in self.config.SUPPORTED_CSV_FORMATS:
                return [self._parse_csv()]
            elif ext in self.config.SUPPORTED_EXCEL_FORMATS:
                if ext in ['.xls', '.xlt']:
                    return self._parse_xls()
                elif ext == '.xlsb':
                    return self._parse_xlsb()
                else:
                    return self._parse_excel()
            elif ext in self.config.SUPPORTED_WPS_FORMATS:
                # WPS格式专门解析
                return self._parse_wps()
            else:
                raise UnsupportedFormatError(self.file_path, ext, context=context)
                
        except Exception as e:
            # 使用统一错误处理器
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error

    def _parse_csv(self):
        """解析CSV文件"""
        context = create_error_context("CSV解析", file_path=self.file_path)
        
        try:
            # 尝试自动检测编码
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
            df = None
            
            for encoding in encodings:
                try:
                    df = pd.read_csv(self.file_path, header=None, dtype=str, encoding=encoding)
                    self.logger.info(f"成功使用编码 {encoding} 读取CSV文件")
                    break
                except UnicodeDecodeError:
                    continue
            
            if df is None:
                raise ParsingError("无法检测CSV文件编码", context=context)
            
            # 清理数据
            data = df.fillna('').astype(str).values.tolist()
            data = [[clean_cell_value(cell) for cell in row] for row in data]
            
        except EmptyDataError:
            data = []
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
        
        # 检查数据大小限制
        self._validate_data_size(data, context)
        
        # CSV无样式
        styles = [[{} for _ in row] for row in data]
        return {
            'sheet_name': os.path.basename(self.file_path),
            'rows': len(data),
            'cols': len(data[0]) if data else 0,
            'data': data,
            'styles': styles,
            'merged_cells': [],  # CSV无合并单元格
            'comments': {},      # CSV无注释
            'hyperlinks': {}     # CSV无超链接
        }

    def _validate_data_size(self, data, context):
        """验证数据大小是否超限"""
        if data and len(data) > self.config.MAX_ROWS:
            context.additional_info = {'rows': len(data), 'max_rows': self.config.MAX_ROWS}
            raise ParsingError(
                f"行数超过限制 ({len(data)} > {self.config.MAX_ROWS})",
                severity=ErrorSeverity.HIGH,
                context=context
            )
            
        if data and len(data[0]) > self.config.MAX_COLS:
            context.additional_info = {'cols': len(data[0]), 'max_cols': self.config.MAX_COLS}
            raise ParsingError(
                f"列数超过限制 ({len(data[0])} > {self.config.MAX_COLS})",
                severity=ErrorSeverity.HIGH,
                context=context
            )

    @error_handler(operation="Excel文件解析")
    def _parse_excel(self):
        """解析现代Excel文件 (.xlsx, .xlsm等)"""
        # 获取文件大小，决定是否使用流式处理
        file_size_mb = os.path.getsize(self.file_path) / (1024 * 1024)
        use_streaming = (file_size_mb > 50 and 
                       getattr(self.config, 'ENABLE_PERFORMANCE_MODE', False))
        
        if use_streaming:
            return self._parse_excel_streaming()
        else:
            return self._parse_excel_standard()
    
    def _parse_excel_standard(self):
        """标准Excel解析方法"""
        context = create_error_context("Excel标准解析", file_path=self.file_path)
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=False)  
            sheets = []
            
            for ws in wb.worksheets:
                try:
                    sheet_context = create_error_context(
                        "工作表解析", 
                        file_path=self.file_path, 
                        sheet_name=ws.title
                    )
                    
                    self.logger.info(f"解析工作表: {ws.title}")
                    
                    # 检查工作表大小
                    self._validate_worksheet_size(ws, sheet_context)
                    
                    # 解析工作表数据
                    sheet_result = self._parse_worksheet_data(ws, sheet_context)
                    
                    # 处理公式（如果启用）
                    if self.formula_processor:
                        sheet_result = safe_execute(
                            self.formula_processor.process_sheet_formulas,
                            sheet_result,
                            operation=f"公式处理-{ws.title}",
                            default_value=sheet_result
                        )
                        self.logger.info(f"工作表 {ws.title} 公式处理完成")
                    
                    # 处理图表（如果启用）
                    if self.chart_converter:
                        sheet_result = safe_execute(
                            self.chart_converter.process_sheet_charts,
                            sheet_result,
                            operation=f"图表转换-{ws.title}",
                            default_value=sheet_result
                        )
                        self.logger.info(f"工作表 {ws.title} 图表处理完成")
                    
                    sheets.append(sheet_result)
                    
                except Exception as e:
                    # 工作表级别的错误处理
                    error = WorksheetParsingError(ws.title, str(e), context=sheet_context, original_error=e)
                    self.error_handler._log_error(error)
                    # 继续处理其他工作表，不中断整体流程
                    continue
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _validate_worksheet_size(self, ws, context):
        """验证工作表大小"""
        if ws.max_row > self.config.MAX_ROWS:
            context.additional_info = {'rows': ws.max_row, 'max_rows': self.config.MAX_ROWS}
            raise ParsingError(
                f"工作表行数超过限制 ({ws.max_row} > {self.config.MAX_ROWS})",
                severity=ErrorSeverity.HIGH,
                context=context
            )
            
        if ws.max_column > self.config.MAX_COLS:
            context.additional_info = {'cols': ws.max_column, 'max_cols': self.config.MAX_COLS}
            raise ParsingError(
                f"工作表列数超过限制 ({ws.max_column} > {self.config.MAX_COLS})",
                severity=ErrorSeverity.HIGH,
                context=context
            )
    
    def _parse_worksheet_data(self, ws, context):
        """解析工作表数据"""
        data = []
        styles = []
        comments = {}  # 注释字典
        hyperlinks = {}  # 超链接字典
        
        # 使用max_row和max_column确保遍历所有行和列
        for r in range(1, ws.max_row + 1):
            data_row = []
            style_row = []
            for c in range(1, ws.max_column + 1):
                try:
                    cell = ws.cell(row=r, column=c)
                    
                    # 获取单元格值，优先获取公式
                    if cell.data_type == 'f':  # 公式单元格
                        # 检查公式是否已经包含等号
                        formula_value = str(cell.value) if cell.value else ''
                        if formula_value and not formula_value.startswith('='):
                            cell_value = f"={formula_value}"
                        else:
                            cell_value = formula_value
                    else:
                        # 对于非公式单元格，直接使用值
                        cell_value = cell.value if cell.value is not None else ''
                    
                    data_row.append(clean_cell_value(cell_value))
                    
                    # 获取样式信息 - 使用安全执行
                    style = safe_execute(
                        _get_cell_style, 
                        cell,
                        operation=f"样式提取-{r},{c}",
                        default_value={}
                    )
                    style_row.append(style)
                    
                    # 提取注释和超链接到独立字典
                    cell_key = f"{r-1}_{c-1}"  # 转换为0-based索引
                    
                    if 'comment' in style and style['comment']:
                        comments[cell_key] = style['comment']
                    
                    if 'hyperlink' in style and style['hyperlink']:
                        hyperlinks[cell_key] = style['hyperlink']
                        
                except Exception as e:
                    # 单元格级别的错误处理
                    cell_pos = f"{r},{c}"
                    error = CellProcessingError(cell_pos, str(e), context=context, original_error=e)
                    self.error_handler._log_error(error)
                    
                    # 使用默认值继续处理
                    data_row.append('')
                    style_row.append({})
                        
            data.append(data_row)
            styles.append(style_row)
        
        # 处理合并单元格
        merged_cells = []
        for merged in ws.merged_cells.ranges:
            merged_cells.append((
                merged.min_row-1, merged.min_col-1, 
                merged.max_row-1, merged.max_col-1
            ))
        
        return {
            'sheet_name': ws.title,
            'rows': len(data),
            'cols': len(data[0]) if data else 0,
            'data': data,
            'styles': styles,
            'merged_cells': merged_cells,
            'comments': comments,
            'hyperlinks': hyperlinks
        }
    
    def _parse_excel_streaming(self):
        """流式Excel解析方法"""
        context = create_error_context("Excel流式解析", file_path=self.file_path)
        
        try:
            self.logger.info("使用流式解析模式处理大文件")
            
            # 创建流式解析器
            chunk_size = getattr(self.config, 'CHUNK_SIZE', 1000)
            streaming_parser = StreamingExcelParser(
                chunk_size=chunk_size,
                enable_memory_monitor=self.memory_monitor is not None
            )
            
            # 创建进度回调
            progress_callback = None
            if (self.progress_callback and 
                getattr(self.config, 'ENABLE_PROGRESS_TRACKING', True)):
                progress_callback = self.progress_callback
            
            # 收集所有数据块
            worksheets = {}
            
            for chunk in streaming_parser.parse_in_chunks(self.file_path, progress_callback):
                ws_name = chunk['worksheet_name']
                
                if ws_name not in worksheets:
                    worksheets[ws_name] = {
                        'sheet_name': ws_name,
                        'data': [],
                        'styles': [],
                        'comments': {},
                        'hyperlinks': {},
                        'merged_cells': []
                    }
                
                # 合并数据块
                ws_data = worksheets[ws_name]
                ws_data['data'].extend(chunk['data'])
                ws_data['styles'].extend(chunk['styles'])
                ws_data['comments'].update(chunk['comments'])
                ws_data['hyperlinks'].update(chunk['hyperlinks'])
                ws_data['merged_cells'].extend(chunk['merged_cells'])
                
                # 内存检查
                if self.memory_monitor and self.memory_monitor.should_trigger_gc():
                    self.memory_monitor.trigger_gc()
            
            # 转换为最终格式
            sheets = []
            for ws_data in worksheets.values():
                ws_data['rows'] = len(ws_data['data'])
                ws_data['cols'] = len(ws_data['data'][0]) if ws_data['data'] else 0
                sheets.append(ws_data)
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error

    @error_handler(operation="XLS文件解析")
    def _parse_xls(self):
        """解析旧版Excel文件 (.xls)"""
        context = create_error_context("XLS解析", file_path=self.file_path)
        
        workbook = xlrd.open_workbook(self.file_path)
        sheets = []
        
        for sheet_index in range(workbook.nsheets):
            worksheet = workbook.sheet_by_index(sheet_index)
            sheet_name = worksheet.name
            
            sheet_context = create_error_context(
                "XLS工作表解析", 
                file_path=self.file_path, 
                sheet_name=sheet_name
            )
            
            try:
                self.logger.info(f"解析工作表: {sheet_name}")
                
                # 检查工作表大小
                if worksheet.nrows > self.config.MAX_ROWS:
                    raise ParsingError(
                        f"XLS工作表行数超过限制 ({worksheet.nrows} > {self.config.MAX_ROWS})",
                        severity=ErrorSeverity.HIGH,
                        context=sheet_context
                    )
                    
                if worksheet.ncols > self.config.MAX_COLS:
                    raise ParsingError(
                        f"XLS工作表列数超过限制 ({worksheet.ncols} > {self.config.MAX_COLS})",
                        severity=ErrorSeverity.HIGH,
                        context=sheet_context
                    )
                
                data = []
                styles = []  # XLS文件样式提取较复杂，暂时用空样式
                
                for row_idx in range(worksheet.nrows):
                    data_row = []
                    style_row = []
                    for col_idx in range(worksheet.ncols):
                        try:
                            cell_value = worksheet.cell_value(row_idx, col_idx)
                            # 处理日期
                            if worksheet.cell_type(row_idx, col_idx) == xlrd.XL_CELL_DATE:
                                if isinstance(cell_value, (int, float)):
                                    date_tuple = xlrd.xldate_as_tuple(cell_value, workbook.datemode)
                                    cell_value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                            data_row.append(clean_cell_value(cell_value))
                            style_row.append({})  # 暂时没有样式信息
                        except IndexError:
                            data_row.append('')
                            style_row.append({})
                    data.append(data_row)
                    styles.append(style_row)
                
                # XLS文件的合并单元格提取较复杂，暂时为空
                merged_cells = []
                
                sheets.append({
                    'sheet_name': sheet_name,
                    'rows': len(data),
                    'cols': len(data[0]) if data else 0,
                    'data': data,
                    'styles': styles,
                    'merged_cells': merged_cells,
                    'comments': {},      # XLS注释提取复杂，暂时为空
                    'hyperlinks': {}     # XLS超链接提取复杂，暂时为空
                })
                
            except Exception as e:
                error = WorksheetParsingError(sheet_name, str(e), context=sheet_context, original_error=e)
                self.error_handler._log_error(error)
                # 继续处理其他工作表
                continue
                
        return sheets
    
    # --------------------------------------------------
    # XLSB 解析实现 (使用 pyxlsb)
    # --------------------------------------------------
    @error_handler(operation="XLSB文件解析")
    def _parse_xlsb(self):
        """解析Excel二进制工作簿 (.xlsb)"""
        context = create_error_context("XLSB解析", file_path=self.file_path)

        try:
            from pyxlsb import open_workbook  # 内部延迟导入，避免未安装时报错
        except ImportError as imp_err:
            raise ParsingError(
                "缺少 pyxlsb 依赖，请运行 `pip install pyxlsb`", 
                severity=ErrorSeverity.HIGH,
                context=context,
                original_error=imp_err
            )

        sheets = []
        try:
            with open_workbook(self.file_path) as wb:
                for sheet_name in wb.sheets:
                    ws = wb.get_sheet(sheet_name)
                    
                    sheet_context = create_error_context(
                        "XLSB工作表解析", 
                        file_path=self.file_path, 
                        sheet_name=sheet_name
                    )
                    
                    try:
                        self.logger.info(f"解析XLSB工作表: {sheet_name}")
                        
                        data = []
                        styles = []
                        comments = {}
                        hyperlinks = {}
                        # 合并单元格支持
                        merged_cells = []
                        # 优先检测ws.merged_cells
                        if hasattr(ws, 'merged_cells') and ws.merged_cells:
                            try:
                                for merged in ws.merged_cells:
                                    # pyxlsb的merged对象通常有row_first, row_last, col_first, col_last属性
                                    merged_cells.append((
                                        merged.row_first, merged.col_first, merged.row_last, merged.col_last
                                    ))
                            except Exception as e:
                                self.logger.warning(f"提取XLSB合并单元格时出错: {e}")
                        # 兼容部分实现为merged_ranges
                        elif hasattr(ws, 'merged_ranges') and ws.merged_ranges:
                            try:
                                for merged in ws.merged_ranges:
                                    merged_cells.append((
                                        merged.row_first, merged.col_first, merged.row_last, merged.col_last
                                    ))
                            except Exception as e:
                                self.logger.warning(f"提取XLSB合并单元格(merged_ranges)时出错: {e}")
                        # 若都不存在则保持空列表
                        formula_cells = {}
                        
                        # 获取工作表范围
                        max_row = 0
                        max_col = 0
                        
                        # 第一遍遍历获取数据范围
                        for row in ws.rows():
                            row_num = row[0].r if row else 0
                            max_row = max(max_row, row_num)
                            for cell in row:
                                max_col = max(max_col, cell.c)
                        
                        # 第二遍遍历填充数据
                        for row_idx in range(1, max_row + 1):
                            data_row = []
                            style_row = []
                            for col_idx in range(1, max_col + 1):
                                cell_value = ''
                                cell_style = {}
                                cell_formula = None
                                cell_key = f"{row_idx-1}_{col_idx-1}"
                                # 查找当前单元格
                                for row in ws.rows():
                                    if row[0].r == row_idx:
                                        for cell in row:
                                            if cell.c == col_idx:
                                                # 处理单元格值
                                                if cell.v is not None:
                                                    if hasattr(cell, 'f') and cell.f:  # 公式
                                                        cell_value = f"={cell.f}"
                                                        cell_formula = cell_value
                                                    else:
                                                        cell_value = str(cell.v)
                                                # 处理样式信息
                                                if hasattr(cell, 's') and cell.s:
                                                    cell_style = self._extract_xlsb_style(cell.s)
                                                # 处理注释
                                                if hasattr(cell, 'comment') and cell.comment:
                                                    comments[cell_key] = str(cell.comment)
                                                # 处理超链接
                                                if hasattr(cell, 'hyperlink') and cell.hyperlink:
                                                    hyperlinks[cell_key] = str(cell.hyperlink)
                                                break
                                        break
                                data_row.append(clean_cell_value(cell_value))
                                style_row.append(cell_style)
                                # 公式单元格结构化
                                if cell_formula:
                                    formula_cells[cell_key] = {
                                        'row': row_idx-1,
                                        'col': col_idx-1,
                                        'formula': cell_formula,
                                        'formula_type': self._classify_formula(cell_formula),
                                        'dependencies': self._extract_formula_dependencies(cell_formula)
                                    }
                            data.append(data_row)
                            styles.append(style_row)
                        # 检查数据大小
                        self._validate_data_size(data, sheet_context)
                        sheets.append({
                            'sheet_name': sheet_name,
                            'rows': len(data),
                            'cols': len(data[0]) if data else 0,
                            'data': data,
                            'styles': styles,
                            'merged_cells': merged_cells,
                            'comments': comments,
                            'hyperlinks': hyperlinks,
                            'formula_cells': formula_cells
                        })
                    except Exception as e:
                        error = WorksheetParsingError(sheet_name, str(e), context=sheet_context, original_error=e)
                        self.error_handler._log_error(error)
                        continue
            return sheets
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _extract_xlsb_style(self, style_obj):
        """提取XLSB格式的样式信息"""
        style = {}
        
        try:
            # 字体样式
            if hasattr(style_obj, 'font'):
                font = style_obj.font
                if hasattr(font, 'bold') and font.bold:
                    style['bold'] = True
                if hasattr(font, 'italic') and font.italic:
                    style['italic'] = True
                if hasattr(font, 'underline') and font.underline:
                    style['underline'] = True
                if hasattr(font, 'strike') and font.strike:
                    style['strike'] = True
                if hasattr(font, 'size'):
                    style['font_size'] = font.size
                if hasattr(font, 'name'):
                    style['font_name'] = font.name
                if hasattr(font, 'color'):
                    style['font_color'] = f"#{font.color[-6:]}" if font.color else None
            
            # 背景色
            if hasattr(style_obj, 'fill'):
                fill = style_obj.fill
                if hasattr(fill, 'fgColor') and fill.fgColor:
                    style['bg_color'] = f"#{fill.fgColor[-6:]}" if fill.fgColor != '00000000' else None
            
            # 对齐方式
            if hasattr(style_obj, 'alignment'):
                align = style_obj.alignment
                if hasattr(align, 'horizontal'):
                    style['align'] = align.horizontal
                if hasattr(align, 'vertical'):
                    style['valign'] = align.vertical
                if hasattr(align, 'wrapText'):
                    style['wrap_text'] = align.wrapText
            
            # 边框
            if hasattr(style_obj, 'border'):
                border = {}
                border_obj = style_obj.border
                
                for side in ['top', 'bottom', 'left', 'right']:
                    if hasattr(border_obj, side):
                        border_side = getattr(border_obj, side)
                        if border_side and hasattr(border_side, 'style') and border_side.style:
                            border[side] = {
                                'style': border_side.style,
                                'color': f"#{border_side.color[-6:]}" if hasattr(border_side, 'color') and border_side.color else '#000000'
                            }
                
                if border:
                    style['border'] = border
                    
        except Exception as e:
            self.logger.warning(f"提取XLSB样式时出错: {e}")
        
        return style
    
    # --------------------------------------------------
    # WPS 格式解析实现
    # --------------------------------------------------
    @error_handler(operation="WPS文件解析")
    def _parse_wps(self):
        """解析WPS Office格式文件 (.et, .ett, .ets)"""
        context = create_error_context("WPS解析", file_path=self.file_path)
        
        try:
            ext = get_file_extension(self.file_path)
            self.logger.info(f"开始解析WPS文件: {self.file_path} (格式: {ext})")
            
            # WPS格式实际上是基于Excel格式的，但有一些特殊处理
            if ext == '.et':
                return self._parse_wps_workbook()
            elif ext == '.ett':
                return self._parse_wps_template()
            elif ext == '.ets':
                return self._parse_wps_backup()
            else:
                # 默认按Excel处理
                return self._parse_excel()
                
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _parse_wps_workbook(self):
        """解析WPS工作簿文件 (.et)"""
        context = create_error_context("WPS工作簿解析", file_path=self.file_path)
        
        try:
            # 首先尝试使用openpyxl解析
            try:
                sheets = self._parse_excel_standard()
            except Exception as excel_error:
                self.logger.warning(f"openpyxl无法解析.et文件，尝试使用pandas: {excel_error}")
                # 如果openpyxl失败，使用pandas作为备选方案
                sheets = self._parse_wps_with_pandas()
            
            # 添加WPS特有的元数据处理和公式增强
            for sheet in sheets:
                sheet['wps_format'] = 'workbook'
                sheet['wps_metadata'] = self._extract_wps_metadata()
                
                # 增强公式支持
                sheet['formula_cells'] = self._extract_formula_cells(sheet)
                sheet['wps_formula_info'] = self._analyze_wps_formulas(sheet)
                # WPS专有字段映射
                sheet['wps_comments'] = sheet.get('comments', {})
                sheet['wps_hyperlinks'] = sheet.get('hyperlinks', {})
                sheet['wps_merged_cells'] = sheet.get('merged_cells', [])
                sheet['wps_styles'] = sheet.get('styles', [])
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _parse_wps_template(self):
        """解析WPS模板文件 (.ett)"""
        context = create_error_context("WPS模板解析", file_path=self.file_path)
        
        try:
            # 首先尝试使用openpyxl解析
            try:
                sheets = self._parse_excel_standard()
            except Exception as excel_error:
                self.logger.warning(f"openpyxl无法解析.ett文件，尝试使用pandas: {excel_error}")
                # 如果openpyxl失败，使用pandas作为备选方案
                sheets = self._parse_wps_with_pandas()
            
            # 添加模板特有的处理
            for sheet in sheets:
                sheet['wps_format'] = 'template'
                sheet['wps_metadata'] = self._extract_wps_metadata()
                sheet['is_template'] = True
                
                # 处理模板变量（如果有的话）
                sheet['template_variables'] = self._extract_template_variables(sheet)
                
                # 增强公式支持
                sheet['formula_cells'] = self._extract_formula_cells(sheet)
                sheet['wps_formula_info'] = self._analyze_wps_formulas(sheet)
                # WPS专有字段映射
                sheet['wps_comments'] = sheet.get('comments', {})
                sheet['wps_hyperlinks'] = sheet.get('hyperlinks', {})
                sheet['wps_merged_cells'] = sheet.get('merged_cells', [])
                sheet['wps_styles'] = sheet.get('styles', [])
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _parse_wps_with_pandas(self):
        """使用pandas解析WPS文件（当openpyxl不支持时）"""
        context = create_error_context("WPS Pandas解析", file_path=self.file_path)
        
        try:
            self.logger.info("使用pandas解析WPS文件")
            
            # 尝试使用pandas读取所有工作表
            excel_file = pd.ExcelFile(self.file_path)
            sheets = []
            
            for sheet_name in excel_file.sheet_names:
                try:
                    sheet_context = create_error_context(
                        "WPS工作表解析", 
                        file_path=self.file_path, 
                        sheet_name=sheet_name
                    )
                    
                    self.logger.info(f"解析WPS工作表: {sheet_name}")
                    
                    # 读取工作表数据
                    df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
                    
                    # 转换为列表格式
                    data = df.fillna('').astype(str).values.tolist()
                    data = [[clean_cell_value(cell) for cell in row] for row in data]
                    
                    # 检查数据大小限制
                    if len(data) > self.config.MAX_ROWS:
                        raise ParsingError(
                            f"WPS工作表行数超过限制 ({len(data)} > {self.config.MAX_ROWS})",
                            severity=ErrorSeverity.HIGH,
                            context=sheet_context
                        )
                    
                    if data and len(data[0]) > self.config.MAX_COLS:
                        raise ParsingError(
                            f"WPS工作表列数超过限制 ({len(data[0])} > {self.config.MAX_COLS})",
                            severity=ErrorSeverity.HIGH,
                            context=sheet_context
                        )
                    
                    # 创建样式矩阵（pandas无法提取样式，使用默认样式）
                    styles = [[{} for _ in row] for row in data]
                    
                    sheet_result = {
                        'sheet_name': sheet_name,
                        'rows': len(data),
                        'cols': len(data[0]) if data else 0,
                        'data': data,
                        'styles': styles,
                        'merged_cells': [],  # pandas无法提取合并单元格
                        'comments': {},      # pandas无法提取注释
                        'hyperlinks': {}     # pandas无法提取超链接
                    }
                    
                    sheets.append(sheet_result)
                    
                except Exception as e:
                    # 工作表级别的错误处理
                    error = WorksheetParsingError(sheet_name, str(e), context=sheet_context, original_error=e)
                    self.error_handler._log_error(error)
                    # 继续处理其他工作表，不中断整体流程
                    continue
            
            if not sheets:
                raise ParsingError("无法解析WPS文件中的任何工作表", context=context)
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _parse_wps_backup(self):
        """解析WPS备份文件 (.ets)"""
        context = create_error_context("WPS备份解析", file_path=self.file_path)
        
        try:
            # WPS .ets文件是备份格式
            # 首先尝试使用openpyxl解析
            try:
                sheets = self._parse_excel_standard()
            except Exception as excel_error:
                self.logger.warning(f"openpyxl无法解析.ets文件，尝试使用pandas: {excel_error}")
                # 如果openpyxl失败，使用pandas作为备选方案
                sheets = self._parse_wps_with_pandas()
            
            # 添加备份特有的处理
            for sheet in sheets:
                sheet['wps_format'] = 'backup'
                sheet['wps_metadata'] = self._extract_wps_metadata()
                sheet['is_backup'] = True
                
                # 处理备份信息
                sheet['backup_info'] = self._extract_backup_info()
                
                # 增强公式支持
                sheet['formula_cells'] = self._extract_formula_cells(sheet)
                sheet['wps_formula_info'] = self._analyze_wps_formulas(sheet)
                # WPS专有字段映射
                sheet['wps_comments'] = sheet.get('comments', {})
                sheet['wps_hyperlinks'] = sheet.get('hyperlinks', {})
                sheet['wps_merged_cells'] = sheet.get('merged_cells', [])
                sheet['wps_styles'] = sheet.get('styles', [])
            
            return sheets
            
        except Exception as e:
            custom_error = self.error_handler.handle_error(e, context)
            raise custom_error
    
    def _extract_wps_metadata(self):
        """提取WPS文件元数据"""
        metadata = {
            'format_type': 'wps',
            'version': 'unknown',
            'creator': 'WPS Office',
            'created_time': None,
            'modified_time': None
        }
        
        try:
            # 尝试从Excel属性中提取WPS特有信息
            workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            
            # 提取文档属性
            if hasattr(workbook, 'properties'):
                props = workbook.properties
                if hasattr(props, 'created'):
                    metadata['created_time'] = str(props.created)
                if hasattr(props, 'modified'):
                    metadata['modified_time'] = str(props.modified)
                if hasattr(props, 'creator'):
                    metadata['creator'] = props.creator
                if hasattr(props, 'title'):
                    metadata['title'] = props.title
            
            # 尝试从自定义属性中提取WPS版本信息
            if hasattr(workbook, 'custom_document_properties'):
                for prop in workbook.custom_document_properties:
                    if 'wps' in prop.name.lower() or 'version' in prop.name.lower():
                        metadata['version'] = prop.value
                        break
            
            workbook.close()
            
        except Exception as e:
            self.logger.warning(f"提取WPS元数据时出错: {e}")
            # 当openpyxl失败时，尝试从文件系统获取基本信息
            try:
                import os
                from datetime import datetime
                
                # 获取文件修改时间
                file_mtime = os.path.getmtime(self.file_path)
                metadata['modified_time'] = datetime.fromtimestamp(file_mtime).isoformat()
                
                # 获取文件大小
                file_size = os.path.getsize(self.file_path)
                metadata['file_size'] = file_size
                
                # 从文件名推断格式类型
                file_name = os.path.basename(self.file_path)
                ext = get_file_extension(self.file_path)
                if ext == '.ett':
                    metadata['format_type'] = 'wps_template'
                elif ext == '.ets':
                    metadata['format_type'] = 'wps_backup'
                else:
                    metadata['format_type'] = 'wps_workbook'
                    
            except Exception as fs_error:
                self.logger.warning(f"从文件系统获取元数据时出错: {fs_error}")
        
        return metadata
    
    def _extract_template_variables(self, sheet):
        """提取模板变量"""
        template_vars = {}
        
        try:
            # 在数据中查找可能的模板变量（如 {{变量名}} 格式）
            for row_idx, row in enumerate(sheet['data']):
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str) and '{{' in cell_value and '}}' in cell_value:
                        # 简单的模板变量提取
                        import re
                        matches = re.findall(r'\{\{([^}]+)\}\}', cell_value)
                        for match in matches:
                            template_vars[match.strip()] = {
                                'row': row_idx,
                                'col': col_idx,
                                'original_value': cell_value
                            }
            
        except Exception as e:
            self.logger.warning(f"提取模板变量时出错: {e}")
        
        return template_vars
    
    def _extract_backup_info(self):
        """提取备份信息"""
        backup_info = {
            'backup_time': None,
            'original_file': None,
            'backup_reason': 'auto'
        }
        
        try:
            # 尝试从文件名或路径中提取备份信息
            import os
            from datetime import datetime
            
            file_name = os.path.basename(self.file_path)
            file_dir = os.path.dirname(self.file_path)
            
            # 尝试从文件名中提取时间信息
            if '_backup_' in file_name or '_bak_' in file_name:
                backup_info['backup_reason'] = 'manual'
            
            # 尝试从文件修改时间获取备份时间
            file_mtime = os.path.getmtime(self.file_path)
            backup_info['backup_time'] = datetime.fromtimestamp(file_mtime).isoformat()
            
        except Exception as e:
            self.logger.warning(f"提取备份信息时出错: {e}")
        
        return backup_info
    
    def _extract_formula_cells(self, sheet):
        """提取所有公式单元格信息"""
        formula_cells = {}
        
        try:
            for row_idx, row in enumerate(sheet['data']):
                for col_idx, cell_value in enumerate(row):
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        cell_key = f"{row_idx}_{col_idx}"
                        formula_cells[cell_key] = {
                            'row': row_idx,
                            'col': col_idx,
                            'formula': cell_value,
                            'formula_type': self._classify_formula(cell_value),
                            'dependencies': self._extract_formula_dependencies(cell_value)
                        }
            
        except Exception as e:
            self.logger.warning(f"提取公式单元格时出错: {e}")
        
        return formula_cells
    
    def _analyze_wps_formulas(self, sheet):
        """分析WPS公式兼容性"""
        wps_formula_info = {
            'total_formulas': 0,
            'standard_formulas': 0,
            'wps_specific_formulas': 0,
            'compatibility_issues': [],
            'formula_types': {}
        }
        
        try:
            formula_cells = sheet.get('formula_cells', {})
            wps_formula_info['total_formulas'] = len(formula_cells)
            
            for cell_key, formula_info in formula_cells.items():
                formula = formula_info['formula']
                formula_type = formula_info['formula_type']
                
                # 统计公式类型
                if formula_type not in wps_formula_info['formula_types']:
                    wps_formula_info['formula_types'][formula_type] = 0
                wps_formula_info['formula_types'][formula_type] += 1
                
                # 检查是否为WPS特有公式
                if self._is_wps_specific_formula(formula):
                    wps_formula_info['wps_specific_formulas'] += 1
                    wps_formula_info['compatibility_issues'].append({
                        'cell': cell_key,
                        'formula': formula,
                        'issue': 'WPS特有公式，可能在其他软件中不兼容'
                    })
                else:
                    wps_formula_info['standard_formulas'] += 1
                
                # 检查其他兼容性问题
                compatibility_issue = self._check_formula_compatibility(formula)
                if compatibility_issue:
                    wps_formula_info['compatibility_issues'].append({
                        'cell': cell_key,
                        'formula': formula,
                        'issue': compatibility_issue
                    })
            
        except Exception as e:
            self.logger.warning(f"分析WPS公式时出错: {e}")
        
        return wps_formula_info
    
    def _classify_formula(self, formula):
        """分类公式类型"""
        formula_upper = formula.upper()
        
        # 数学函数
        if any(func in formula_upper for func in ['SUM(', 'AVERAGE(', 'COUNT(', 'MAX(', 'MIN(', 'ROUND(', 'ABS(']):
            return 'mathematical'
        
        # 文本函数
        elif any(func in formula_upper for func in ['CONCATENATE(', 'LEFT(', 'RIGHT(', 'MID(', 'LEN(', 'UPPER(', 'LOWER(']):
            return 'text'
        
        # 逻辑函数
        elif any(func in formula_upper for func in ['IF(', 'AND(', 'OR(', 'NOT(', 'TRUE(', 'FALSE(']):
            return 'logical'
        
        # 查找函数
        elif any(func in formula_upper for func in ['VLOOKUP(', 'HLOOKUP(', 'INDEX(', 'MATCH(']):
            return 'lookup'
        
        # 日期时间函数
        elif any(func in formula_upper for func in ['TODAY(', 'NOW(', 'DATE(', 'YEAR(', 'MONTH(', 'DAY(']):
            return 'datetime'
        
        # 引用函数
        elif any(func in formula_upper for func in ['INDIRECT(', 'OFFSET(', 'ADDRESS(']):
            return 'reference'
        
        # 简单计算
        elif any(op in formula for op in ['+', '-', '*', '/', '^']):
            return 'calculation'
        
        # 单元格引用
        elif any(char in formula for char in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']):
            return 'reference'
        
        else:
            return 'unknown'
    
    def _extract_formula_dependencies(self, formula):
        """提取公式依赖的单元格"""
        dependencies = []
        try:
            import re
            # 支持 Sheet1!A1, Sheet1!A:C, Sheet1!A1:B5, Sheet1!A:B
            sheet_ref_pattern = r'([A-Z0-9_]+!([A-Z]+\d+(:[A-Z]+\d+)?|[A-Z]+(:[A-Z]+)?))'
            matches = re.findall(sheet_ref_pattern, formula.upper())
            for m in matches:
                dependencies.append(m[0])
            # 范围引用 A1:B5
            range_pattern = r'\b[A-Z]+\d+:[A-Z]+\d+\b'
            range_matches = re.findall(range_pattern, formula.upper())
            dependencies.extend(range_matches)
            # 列范围引用 A:B
            col_range_pattern = r'\b[A-Z]+:[A-Z]+\b'
            col_range_matches = re.findall(col_range_pattern, formula.upper())
            dependencies.extend(col_range_matches)
            # 简单单元格引用 A1, B2
            cell_pattern = r'\b[A-Z]+\d+\b'
            cell_matches = re.findall(cell_pattern, formula.upper())
            # 过滤掉已在Sheet1!A1中的
            for cell in cell_matches:
                if not any(cell in dep for dep in dependencies if '!' in dep):
                    dependencies.append(cell)
            dependencies = list(set(dependencies))
        except Exception as e:
            self.logger.warning(f"提取公式依赖时出错: {e}")
        return dependencies
    
    def _is_wps_specific_formula(self, formula):
        """检查是否为WPS特有公式"""
        # WPS可能有一些特有的函数或语法模式
        wps_specific_patterns = [
            # 这里可以添加WPS特有的函数名或语法模式
            # 例如：'WPS_', 'KINGSOFT_', 等
        ]
        
        formula_upper = formula.upper()
        for pattern in wps_specific_patterns:
            if pattern in formula_upper:
                return True
        
        return False
    
    def _check_formula_compatibility(self, formula):
        """检查公式兼容性"""
        compatibility_issues = []
        
        try:
            # 检查是否包含可能不兼容的函数
            problematic_functions = [
                # 可以添加已知的兼容性问题函数
            ]
            
            formula_upper = formula.upper()
            for func in problematic_functions:
                if func in formula_upper:
                    compatibility_issues.append(f"包含可能不兼容的函数: {func}")
            
            # 检查公式长度
            if len(formula) > 8192:  # Excel公式长度限制
                compatibility_issues.append("公式长度超过Excel限制")
            
            # 检查嵌套层级
            open_parens = formula.count('(')
            close_parens = formula.count(')')
            if open_parens != close_parens:
                compatibility_issues.append("括号不匹配")
            
        except Exception as e:
            self.logger.warning(f"检查公式兼容性时出错: {e}")
        
        return '; '.join(compatibility_issues) if compatibility_issues else None
    
    def get_error_summary(self):
        """获取解析过程中的错误统计"""
        return self.error_handler.get_error_summary() 