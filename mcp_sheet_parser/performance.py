# performance.py
# 性能优化模块 - 大文件处理、进度跟踪、内存优化

import gc
import time
import psutil
import threading
import logging
from typing import Iterator, Callable, Optional, Dict, Any, List, Tuple
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.utils import get_column_letter

from .utils import clean_cell_value


class ProgressTracker:
    """进度跟踪器"""
    
    def __init__(self, total_size: int, description: str = "处理中"):
        self.total_size = total_size
        self.processed = 0
        self.start_time = time.time()
        self.description = description
        self.callbacks: List[Callable] = []
        self.last_update_time = 0
        self.update_interval = 0.1  # 100ms更新间隔
        
    def add_callback(self, callback: Callable[[Dict[str, Any]], None]):
        """添加进度回调函数"""
        self.callbacks.append(callback)
    
    def update(self, increment: int = 1):
        """更新进度"""
        self.processed += increment
        current_time = time.time()
        
        # 限制更新频率
        if current_time - self.last_update_time < self.update_interval:
            return
        
        self.last_update_time = current_time
        progress_info = self._calculate_progress()
        
        # 调用所有回调函数
        for callback in self.callbacks:
            try:
                callback(progress_info)
            except Exception as e:
                logging.warning(f"进度回调函数执行失败: {e}")
    
    def _calculate_progress(self) -> Dict[str, Any]:
        """计算进度信息"""
        if self.total_size == 0:
            percentage = 100.0
        else:
            percentage = (self.processed / self.total_size) * 100
        
        elapsed_time = time.time() - self.start_time
        if self.processed > 0 and elapsed_time > 0:
            rate = self.processed / elapsed_time
            remaining = (self.total_size - self.processed) / rate if rate > 0 else 0
        else:
            rate = 0
            remaining = 0
        
        # 计算预计完成时间
        if rate > 0:
            eta_seconds = remaining / rate
            eta = f"{eta_seconds:.1f}秒"
        else:
            eta = "未知"
        
        return {
            'percentage': min(percentage, 100.0),
            'processed': self.processed,
            'total': self.total_size,
            'elapsed': elapsed_time,
            'remaining': remaining,
            'rate': rate,
            'speed': f"{rate:.1f} 项/秒",
            'eta': eta,  # 添加eta字段
            'description': self.description
        }
    
    def is_complete(self) -> bool:
        """检查是否完成"""
        return self.processed >= self.total_size
    
    def finish(self):
        """完成进度跟踪"""
        self.processed = self.total_size
        final_info = self._calculate_progress()
        for callback in self.callbacks:
            try:
                callback(final_info)
            except Exception as e:
                logging.warning(f"进度回调函数执行失败: {e}")


class MemoryMonitor:
    """内存监控器"""
    
    def __init__(self, max_memory_mb: int = 2048):
        self.max_memory_mb = max_memory_mb
        self.max_memory_bytes = max_memory_mb * 1024 * 1024
        self.process = psutil.Process()
        self.gc_threshold = 0.8  # 80%时触发GC
        
    @property
    def max_memory(self) -> int:
        """最大内存限制（字节）"""
        return self.max_memory_bytes
    
    def get_memory_usage(self) -> Dict[str, Any]:
        """获取当前内存使用情况"""
        memory_info = self.process.memory_info()
        memory_percent = self.process.memory_percent()
        
        return {
            'rss': memory_info.rss,
            'vms': memory_info.vms,
            'rss_mb': memory_info.rss / (1024 * 1024),
            'percent': memory_percent,
            'max_allowed': self.max_memory_bytes,
            'usage_ratio': memory_info.rss / self.max_memory_bytes
        }
    
    def should_trigger_gc(self) -> bool:
        """检查是否应该触发垃圾回收"""
        usage = self.get_memory_usage()
        return usage['usage_ratio'] > self.gc_threshold
    
    def trigger_gc(self) -> float:
        """触发垃圾回收"""
        before = self.get_memory_usage()
        
        # 执行垃圾回收
        collected = gc.collect()
        
        after = self.get_memory_usage()
        
        freed_memory = float(before['rss'] - after['rss'])  # 确保返回float
        return freed_memory
    
    def get_memory_stats(self) -> str:
        """获取内存统计信息的字符串表示"""
        stats = self.get_memory_usage()
        return f"内存使用: {stats['rss'] / 1024 / 1024:.1f}MB ({stats['percent']:.1f}%)"


class PerformanceOptimizer:
    """性能优化器"""
    
    def __init__(self):
        self.optimization_cache = {}
        
    def get_system_info(self) -> Dict[str, Any]:
        """获取系统信息"""
        return {
            'cpu_count': psutil.cpu_count() or 1,
            'memory_total_gb': psutil.virtual_memory().total / (1024**3),
            'memory_available_gb': psutil.virtual_memory().available / (1024**3),
            'memory_percent': psutil.virtual_memory().percent
        }
        
    def optimize_for_file(self, file_path: str) -> 'PerformanceConfig':
        """根据文件特征优化性能配置"""
        from .config.performance import PerformanceConfig
        
        file_size = Path(file_path).stat().st_size
        file_size_mb = file_size / (1024 * 1024)
        
        # 创建配置对象
        config = PerformanceConfig()
        
        # 基于文件大小的优化策略
        if file_size_mb < 10:
            # 小文件 - 优先处理速度
            config.CHUNK_SIZE = 2000
            config.ENABLE_PARALLEL_PROCESSING = False
            config.MAX_MEMORY_MB = 512
        elif file_size_mb < 100:
            # 中等文件 - 平衡速度和内存
            config.CHUNK_SIZE = 1000
            config.ENABLE_PARALLEL_PROCESSING = True
            config.MAX_MEMORY_MB = 1024
        else:
            # 大文件 - 优先内存效率
            config.CHUNK_SIZE = 500
            config.ENABLE_PARALLEL_PROCESSING = True
            config.MAX_MEMORY_MB = 2048
        
        return config


class StreamingExcelParser:
    """流式Excel解析器"""
    
    def __init__(self, chunk_size: int = 1000, enable_memory_monitor: bool = True):
        self.chunk_size = chunk_size
        self.memory_monitor = MemoryMonitor() if enable_memory_monitor else None
        
    def _extract_chunk_data(self, worksheet, start_row: int, end_row: int) -> Dict[str, Any]:
        """提取数据块"""
        data = []
        styles = []
        comments = {}
        hyperlinks = {}
        
        for row_idx in range(start_row, min(end_row, worksheet.max_row + 1)):
            data_row = []
            style_row = []
            
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # 数据
                cell_value = cell.value if cell.value is not None else ''
                data_row.append(clean_cell_value(cell_value))
                
                # 样式（简化）
                style_row.append({})
                
                # 注释和超链接（简化）
                if cell.comment:
                    comments[f"{row_idx-1}_{col_idx-1}"] = cell.comment.text
                if cell.hyperlink:
                    hyperlinks[f"{row_idx-1}_{col_idx-1}"] = str(cell.hyperlink.target)
            
            data.append(data_row)
            styles.append(style_row)
        
        return {
            'data': data,
            'styles': styles,
            'comments': comments,
            'hyperlinks': hyperlinks,
            'merged_cells': []
        }
    
    def parse_in_chunks(self, file_path: str, progress_callback: Optional[Callable] = None) -> Iterator[Dict[str, Any]]:
        """
        分块解析Excel文件
        
        Args:
            file_path: Excel文件路径
            progress_callback: 进度回调函数
            
        Yields:
            数据块字典
        """
        from .parser import _get_cell_style
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        for ws in wb.worksheets:
            total_rows = ws.max_row
            progress_tracker = None
            
            if progress_callback:
                progress_tracker = ProgressTracker(total_rows, f"解析工作表 {ws.title}")
                progress_tracker.add_callback(progress_callback)
            
            current_chunk = []
            current_styles = []
            row_count = 0
            
            for row in ws.iter_rows():
                data_row = []
                style_row = []
                
                for cell in row:
                    cell_value = cell.value if cell.value is not None else ''
                    data_row.append(clean_cell_value(cell_value))
                    
                    # 简化的样式提取（read_only模式限制）
                    style_row.append({})
                
                current_chunk.append(data_row)
                current_styles.append(style_row)
                row_count += 1
                
                # 当达到块大小时，产出数据块
                if len(current_chunk) >= self.chunk_size:
                    yield {
                        'worksheet_name': ws.title,
                        'data': current_chunk,
                        'styles': current_styles,
                        'comments': {},  # read_only模式下难以获取
                        'hyperlinks': {},  # read_only模式下难以获取
                        'merged_cells': [],  # read_only模式下难以获取
                        'chunk_info': {
                            'start_row': row_count - len(current_chunk),
                            'end_row': row_count - 1,
                            'chunk_size': len(current_chunk)
                        }
                    }
                    
                    # 重置块数据
                    current_chunk = []
                    current_styles = []
                
                # 更新进度
                if progress_tracker:
                    progress_tracker.update()
                
                # 内存检查
                if self.memory_monitor and self.memory_monitor.should_trigger_gc():
                    gc_stats = self.memory_monitor.trigger_gc()
                    logging.info(f"触发垃圾回收: 释放 {gc_stats / 1024 / 1024:.1f}MB")
            
            # 处理最后的不完整块
            if current_chunk:
                yield {
                    'worksheet_name': ws.title,
                    'data': current_chunk,
                    'styles': current_styles,
                    'comments': {},
                    'hyperlinks': {},
                    'merged_cells': [],
                    'chunk_info': {
                        'start_row': row_count - len(current_chunk),
                        'end_row': row_count - 1,
                        'chunk_size': len(current_chunk)
                    }
                }
            
            # 完成进度跟踪
            if progress_tracker:
                progress_tracker.finish()


class ParallelProcessor:
    """并行处理器"""
    
    def __init__(self, max_workers: int = None):
        self.max_workers = max_workers or min(4, (psutil.cpu_count() or 1) + 1)
        
    def process_worksheets_parallel(self, worksheets: List[Any], process_func: Callable) -> List[Dict[str, Any]]:
        """并行处理工作表"""
        results = []
        
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_ws = {
                executor.submit(process_func, ws): ws 
                for ws in worksheets
            }
            
            # 收集结果
            for future in as_completed(future_to_ws):
                try:
                    result = future.result()
                    results.append(result)
                except Exception as e:
                    ws = future_to_ws[future]
                    logging.error(f"处理工作表失败 {ws}: {e}")
                    results.append({'error': str(e), 'worksheet': ws})
        
        return results


class PerformanceBenchmark:
    """性能基准测试"""
    
    def __init__(self):
        self.results = {}
        
    def benchmark_parse_time(self, file_path: str) -> float:
        """测试解析时间"""
        from .parser import SheetParser
        from .config import Config
        
        start_time = time.time()
        
        config = Config()
        parser = SheetParser(file_path, config)
        sheets = parser.parse()
        
        end_time = time.time()
        return end_time - start_time
    
    def benchmark_memory_usage(self, file_path: str) -> Dict[str, Any]:
        """测试内存使用"""
        memory_monitor = MemoryMonitor()
        
        before = memory_monitor.get_memory_usage()
        
        # 执行解析
        parse_time = self.benchmark_parse_time(file_path)
        
        after = memory_monitor.get_memory_usage()
        
        return {
            'parse_time': parse_time,
            'memory_before': before['rss'] / 1024 / 1024,  # MB
            'memory_after': after['rss'] / 1024 / 1024,   # MB
            'memory_peak': after['rss'] / 1024 / 1024,    # MB (简化)
            'memory_increase': (after['rss'] - before['rss']) / 1024 / 1024
        }
    
    def run_comprehensive_benchmark(self, file_path: str) -> Dict[str, Any]:
        """运行综合基准测试"""
        file_size = Path(file_path).stat().st_size / 1024 / 1024  # MB
        
        # 解析性能测试
        memory_stats = self.benchmark_memory_usage(file_path)
        
        return {
            'file_size_mb': f"{file_size:.2f}",
            'parse_time_seconds': f"{memory_stats['parse_time']:.2f}",
            'memory_usage_mb': f"{memory_stats['memory_increase']:.2f}",
            'throughput_mb_per_sec': f"{file_size / memory_stats['parse_time']:.2f}" if memory_stats['parse_time'] > 0 else "N/A"
        }


def create_progress_callback(verbose: bool = False) -> Callable:
    """创建进度回调函数"""
    def callback(progress_info: Dict[str, Any]):
        if verbose:
            percentage = progress_info.get('percentage', 0)
            description = progress_info.get('description', '')
            print(f"\r{description}: {percentage:.1f}%", end='', flush=True)
            
            if percentage >= 100:
                print()  # 完成时换行
        else:
            # 简单的进度显示
            percentage = progress_info.get('percentage', 0)
            print(f"\r进度: {percentage:.1f}%", end='', flush=True)
            
            if percentage >= 100:
                print()
    
    return callback


def benchmark_performance(func: Callable, *args, **kwargs) -> Dict[str, Any]:
    """基准测试函数性能"""
    import time
    
    start_time = time.time()
    start_memory = psutil.Process().memory_info().rss
    
    try:
        result = func(*args, **kwargs)
        success = True
        error = None
    except Exception as e:
        result = None
        success = False
        error = str(e)
    
    end_time = time.time()
    end_memory = psutil.Process().memory_info().rss
    
    execution_time = end_time - start_time
    memory_used = end_memory - start_memory
    
    return {
        'success': success,
        'execution_time': execution_time,
        'memory_used_mb': memory_used / (1024 * 1024),
        'memory_delta_mb': memory_used / (1024 * 1024),
        'peak_memory_mb': max(start_memory, end_memory) / (1024 * 1024),  # 添加peak_memory_mb字段
        'result': result,
        'error': error
    } 