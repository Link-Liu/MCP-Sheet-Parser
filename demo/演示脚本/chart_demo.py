#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图表转换演示脚本
专门用于展示MCP-Sheet-Parser的图表转换功能
"""

import os
import sys
import subprocess
import time
import webbrowser
from pathlib import Path

# 修复Windows编码问题
if os.name == 'nt':
    try:
        os.system('chcp 65001 > nul')
        os.environ['PYTHONIOENCODING'] = 'utf-8'
        os.environ['PYTHONLEGACYWINDOWSSTDIO'] = 'utf-8'
    except Exception:
        pass

def print_header(title):
    """打印标题"""
    print("\n" + "=" * 60)
    print(f"📊 {title}")
    print("=" * 60)

def print_step(step, description):
    """打印步骤信息"""
    print(f"\n📋 步骤 {step}: {description}")
    print("-" * 40)

def run_command(command, description=""):
    """运行命令并显示结果"""
    if description:
        print(f"🔄 {description}")
    
    try:
        env = os.environ.copy()
        if os.name == 'nt':
            env['PYTHONIOENCODING'] = 'utf-8'
            env['PYTHONLEGACYWINDOWSSTDIO'] = 'utf-8'
        
        result = subprocess.run(
            command, 
            shell=True, 
            capture_output=True, 
            encoding='utf-8', 
            errors='replace',
            env=env
        )
        
        stdout = result.stdout.strip() if result.stdout else ''
        stderr = result.stderr.strip() if result.stderr else ''
        
        if result.returncode == 0:
            print(f"✅ 成功: {description}")
            if stdout:
                lines = stdout.split('\n')
                filtered_lines = []
                for line in lines:
                    if not any(keyword in line for keyword in ['进度:', '|', '处理中', '转换中']):
                        filtered_lines.append(line)
                if filtered_lines:
                    print(f"输出: {' '.join(filtered_lines)}")
        else:
            print(f"❌ 失败: {description}")
            if stderr:
                print(f"错误: {stderr}")
            return False
            
    except Exception as e:
        print(f"❌ 执行出错: {e}")
        return False
    
    return True

def create_chart_sample():
    """创建包含图表的示例文件"""
    print_step(1, "创建图表示例文件")
    
    try:
        import openpyxl
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        
        # 第一个工作表：销售数据
        ws1 = wb.active
        ws1.title = "销售数据"
        
        # 添加数据
        data = [
            ["季度", "产品A", "产品B", "产品C"],
            ["Q1", 120, 100, 80],
            ["Q2", 150, 130, 110],
            ["Q3", 180, 160, 140],
            ["Q4", 200, 170, 150]
        ]
        
        for row in data:
            ws1.append(row)
        
        # 创建柱状图
        chart1 = BarChart()
        chart1.title = "季度销售业绩对比"
        chart1.x_axis.title = "季度"
        chart1.y_axis.title = "销售额(万元)"
        
        data = Reference(ws1, min_col=2, min_row=1, max_col=4, max_row=5)
        cats = Reference(ws1, min_col=1, min_row=2, max_row=5)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        
        ws1.add_chart(chart1, "A7")
        
        # 第二个工作表：市场份额
        ws2 = wb.create_sheet("市场份额")
        
        pie_data = [
            ["地区", "市场份额"],
            ["华东", 35],
            ["华南", 25],
            ["华北", 20],
            ["华中", 15],
            ["其他", 5]
        ]
        
        for row in pie_data:
            ws2.append(row)
        
        # 创建饼图
        chart2 = PieChart()
        chart2.title = "市场份额分布"
        
        data = Reference(ws2, min_col=2, min_row=1, max_row=6)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=6)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        
        ws2.add_chart(chart2, "A8")
        
        # 第三个工作表：增长趋势
        ws3 = wb.create_sheet("增长趋势")
        
        trend_data = [
            ["月份", "收入", "利润"],
            ["1月", 100, 20],
            ["2月", 120, 25],
            ["3月", 110, 22],
            ["4月", 140, 30],
            ["5月", 160, 35],
            ["6月", 180, 40]
        ]
        
        for row in trend_data:
            ws3.append(row)
        
        # 创建折线图
        chart3 = LineChart()
        chart3.title = "月度增长趋势"
        chart3.x_axis.title = "月份"
        chart3.y_axis.title = "金额(万元)"
        
        data = Reference(ws3, min_col=2, min_row=1, max_col=3, max_row=7)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=7)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        
        ws3.add_chart(chart3, "A9")
        
        # 保存文件
        os.makedirs("示例文件", exist_ok=True)
        wb.save("示例文件/chart_demo.xlsx")
        print("✅ 创建图表示例文件: chart_demo.xlsx")
        return True
        
    except ImportError:
        print("❌ 需要安装openpyxl: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ 创建图表示例文件失败: {e}")
        return False

def run_chart_conversions():
    """运行图表转换演示"""
    print_step(2, "图表转换演示")
    
    chart_demos = [
        {
            'input': '示例文件/chart_demo.xlsx',
            'output': 'demo/静态展示/chart_basic.html',
            'description': '基础图表转换（默认设置）',
            'options': ''
        },
        {
            'input': '示例文件/chart_demo.xlsx',
            'output': 'demo/静态展示/chart_high_quality.html',
            'description': '高质量图表转换',
            'options': '--chart-quality high --chart-width 800 --chart-height 500'
        },
        {
            'input': '示例文件/chart_demo.xlsx',
            'output': 'demo/静态展示/chart_responsive.html',
            'description': '响应式图表转换',
            'options': '--chart-responsive --chart-format svg'
        },
        {
            'input': '示例文件/chart_demo.xlsx',
            'output': 'demo/静态展示/chart_minimal.html',
            'description': '极简主题图表',
            'options': '--theme minimal --chart-quality low'
        },
        {
            'input': '示例文件/chart_demo.xlsx',
            'output': 'demo/静态展示/chart_dark.html',
            'description': '暗色主题图表',
            'options': '--theme dark --chart-format svg'
        }
    ]
    
    success_count = 0
    for demo in chart_demos:
        if os.path.exists(demo['input']):
            command = f"python main.py {demo['input']} -o {demo['output']} {demo['options']}"
            
            if run_command(command, demo['description']):
                success_count += 1
        else:
            print(f"⚠️ 跳过 {demo['description']}（文件不存在: {demo['input']}）")
    
    print(f"\n📊 图表转换演示完成: {success_count}/{len(chart_demos)} 成功")
    return success_count > 0

def create_chart_summary():
    """创建图表演示总结"""
    print_step(3, "创建图表演示总结")
    
    summary_content = """# 图表转换功能演示总结

## 📊 图表转换功能概述

MCP-Sheet-Parser提供了强大的图表转换功能，可以将Excel中的图表转换为高质量的SVG矢量图形。

### 🎯 支持的图表类型

- **柱状图 (Bar Chart)**: 适合展示分类数据对比
- **饼图 (Pie Chart)**: 适合展示占比关系
- **折线图 (Line Chart)**: 适合展示趋势变化
- **面积图 (Area Chart)**: 适合展示累积数据
- **散点图 (Scatter Chart)**: 适合展示相关性
- **气泡图 (Bubble Chart)**: 适合展示三维数据
- **环形图 (Donut Chart)**: 饼图的变体
- **雷达图 (Radar Chart)**: 适合展示多维数据
- **仪表图 (Gauge Chart)**: 适合展示进度指标

### ✨ 图表功能特性

#### 1. 输出格式
- **SVG矢量图形**: 高质量、可缩放、文件小
- **PNG位图**: 兼容性好、适合打印

#### 2. 质量选项
- **低质量**: 快速生成，文件小
- **中等质量**: 平衡质量和性能
- **高质量**: 最佳视觉效果

#### 3. 尺寸控制
- 可自定义宽度和高度
- 支持响应式设计
- 自动适应内容

#### 4. 交互功能
- 悬停提示显示数据
- 点击交互（可选）
- 图例交互

#### 5. 主题支持
- 默认主题：专业商务风格
- 暗色主题：深色背景
- 极简主题：简洁清爽
- 自定义主题：可扩展

### 🎨 演示文件说明

本次演示生成了以下图表文件：

1. **chart_basic.html** - 基础图表转换
   - 使用默认设置
   - 展示基本图表功能

2. **chart_high_quality.html** - 高质量图表
   - 800x500像素
   - 高质量渲染
   - 适合正式报告

3. **chart_responsive.html** - 响应式图表
   - SVG格式
   - 自适应屏幕尺寸
   - 适合网页展示

4. **chart_minimal.html** - 极简主题图表
   - 极简主题
   - 低质量设置
   - 快速加载

5. **chart_dark.html** - 暗色主题图表
   - 暗色背景
   - SVG格式
   - 适合夜间模式

### 🚀 使用建议

#### 命令行参数
```bash
# 基础图表转换
python main.py input.xlsx -o output.html

# 高质量图表
python main.py input.xlsx -o output.html --chart-quality high --chart-width 800 --chart-height 500

# 响应式图表
python main.py input.xlsx -o output.html --chart-responsive --chart-format svg

# 主题图表
python main.py input.xlsx -o output.html --theme dark --chart-format svg
```

#### 最佳实践
1. **网页展示**: 使用SVG格式 + 响应式设计
2. **打印输出**: 使用PNG格式 + 高质量设置
3. **性能优化**: 根据需求选择合适的质量级别
4. **主题选择**: 根据使用场景选择合适的主题

### 📈 技术优势

- **矢量图形**: SVG格式保证在任何尺寸下都清晰
- **文件大小**: 相比位图格式，文件更小
- **可编辑性**: SVG格式支持后续编辑
- **兼容性**: 现代浏览器完美支持
- **性能**: 高效的渲染算法

### 🔧 配置选项

图表转换支持以下配置选项：

- `--chart-format`: 输出格式 (svg/png)
- `--chart-quality`: 质量级别 (low/medium/high)
- `--chart-width`: 图表宽度
- `--chart-height`: 图表高度
- `--chart-responsive`: 启用响应式设计
- `--disable-charts`: 禁用图表转换

MCP-Sheet-Parser的图表转换功能已经达到了专业级别，可以满足各种图表展示需求。
"""
    
    summary_file = 'demo/图表演示总结.md'
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(summary_content)
    
    print(f"✅ 图表演示总结已保存到: {summary_file}")
    return True

def open_chart_pages():
    """打开图表演示页面"""
    print_step(4, "打开图表演示页面")
    
    chart_pages = [
        'demo/静态展示/chart_basic.html',
        'demo/静态展示/chart_high_quality.html',
        'demo/静态展示/chart_responsive.html',
        'demo/静态展示/chart_minimal.html',
        'demo/静态展示/chart_dark.html'
    ]
    
    opened_count = 0
    for page in chart_pages:
        if os.path.exists(page):
            try:
                webbrowser.open(f'file://{os.path.abspath(page)}')
                print(f"✅ 已打开: {page}")
                opened_count += 1
                time.sleep(1)
            except Exception as e:
                print(f"❌ 无法打开 {page}: {e}")
    
    print(f"\n📊 已打开 {opened_count} 个图表演示页面")
    return opened_count > 0

def main():
    """主函数"""
    print_header("MCP-Sheet-Parser 图表转换演示")
    
    # 检查当前目录
    if not os.path.exists('main.py'):
        print("❌ 请在项目根目录运行此脚本")
        sys.exit(1)
    
    # 创建必要的目录
    os.makedirs('demo/静态展示', exist_ok=True)
    
    success_count = 0
    total_steps = 4
    
    try:
        # 步骤1: 创建图表示例文件
        if create_chart_sample():
            success_count += 1
        
        # 步骤2: 运行图表转换演示
        if run_chart_conversions():
            success_count += 1
        
        # 步骤3: 创建图表演示总结
        if create_chart_summary():
            success_count += 1
        
        # 步骤4: 打开图表演示页面
        if open_chart_pages():
            success_count += 1
        
        # 显示最终结果
        print_header("图表演示完成")
        print(f"📊 总体进度: {success_count}/{total_steps} 步骤成功")
        
        if success_count == total_steps:
            print("🎉 所有图表演示都成功完成！")
        elif success_count >= total_steps * 0.8:
            print("✅ 大部分图表演示成功完成！")
        else:
            print("⚠️ 部分图表演示失败，请检查错误信息")
        
        print("\n💡 图表功能亮点:")
        print("- 支持10种图表类型转换")
        print("- SVG矢量图形输出")
        print("- 响应式设计支持")
        print("- 多种质量选项")
        print("- 主题化图表样式")
        
    except KeyboardInterrupt:
        print("\n\n⏹️ 用户中断图表演示")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 图表演示过程中出现错误: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 