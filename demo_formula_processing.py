#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
公式处理功能演示程序
展示公式识别、计算和HTML显示功能
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from mcp_sheet_parser.parser import SheetParser
from mcp_sheet_parser.html_converter import HTMLConverter
from mcp_sheet_parser.formula_processor import (
    FormulaProcessor, FormulaCalculator, FormulaType, FormulaError
)
from mcp_sheet_parser.config import Config


def create_formula_demo_excel():
    """创建包含公式的演示Excel文件"""
    print("📊 创建公式演示Excel文件...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "公式演示"
    
    # 设置表头
    headers = ['项目', '单价', '数量', '小计', '折扣率', '实际金额', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 添加数据和公式
    data_rows = [
        ['笔记本电脑', 5000, 3, '=B2*C2', 0.1, '=D2*(1-E2)', '批量采购'],
        ['鼠标', 50, 10, '=B3*C3', 0.05, '=D3*(1-E3)', '办公用品'],
        ['键盘', 200, 8, '=B4*C4', 0.08, '=D4*(1-E4)', '机械键盘'],
        ['显示器', 1200, 5, '=B5*C5', 0.12, '=D5*(1-E5)', '4K显示器'],
        ['', '', '', '', '', '', ''],
        ['统计信息', '', '', '', '', '', ''],
        ['总金额', '', '', '=SUM(D2:D5)', '', '=SUM(F2:F5)', ''],
        ['平均单价', '', '', '=AVERAGE(B2:B5)', '', '', ''],
        ['最高单价', '', '', '=MAX(B2:B5)', '', '', ''],
        ['最低单价', '', '', '=MIN(B2:B5)', '', '', ''],
        ['商品数量', '', '', '=COUNT(B2:B5)', '', '', ''],
        ['', '', '', '', '', '', ''],
        ['数学函数示例', '', '', '', '', '', ''],
        ['平方根', '', '', '=SQRT(16)', '', '', '根号16'],
        ['绝对值', '', '', '=ABS(-25)', '', '', '|-25|'],
        ['四舍五入', '', '', '=ROUND(3.14159,2)', '', '', 'π保留2位'],
        ['幂运算', '', '', '=POWER(2,8)', '', '', '2的8次方'],
        ['', '', '', '', '', '', ''],
        ['逻辑函数示例', '', '', '', '', '', ''],
        ['条件判断', '', '', '=IF(D2>10000,"大订单","小订单")', '', '', ''],
        ['与运算', '', '', '=AND(B2>1000,C2>2)', '', '', '价格>1000且数量>2'],
        ['或运算', '', '', '=OR(E2>0.1,F2>5000)', '', '', '折扣>10%或金额>5000'],
        ['非运算', '', '', '=NOT(B2<100)', '', '', '单价不小于100'],
        ['', '', '', '', '', '', ''],
        ['文本函数示例', '', '', '', '', '', ''],
        ['文本长度', '', '', '=LEN(A2)', '', '', '项目名称长度'],
        ['文本截取', '', '', '=LEFT(A2,2)', '', '', '取前2个字符'],
        ['文本连接', '', '', '=CONCATENATE(A2,"-",G2)', '', '', '项目-备注'],
        ['大写转换', '', '', '=UPPER("hello")', '', '', '转大写'],
        ['小写转换', '', '', '=LOWER("WORLD")', '', '', '转小写'],
        ['', '', '', '', '', '', ''],
        ['错误处理示例', '', '', '', '', '', ''],
        ['除零错误', '', '', '=10/0', '', '', '#DIV/0!'],
        ['引用错误', '', '', '=Z999', '', '', '超出范围'],
        ['函数错误', '', '', '=UNKNOWN(1,2)', '', '', '不支持的函数']
    ]
    
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    output_path = "demo_output/formula_demo.xlsx"
    os.makedirs("demo_output", exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    return output_path


def demo_formula_calculator():
    """演示公式计算器功能"""
    print("\n🧮 === 公式计算器演示 ===")
    
    # 创建测试数据
    sheet_data = {
        'data': [
            ['A', 'B', 'C'],
            ['10', '20', '30'],
            ['15', '25', '35'],
            ['5', '8', '12']
        ]
    }
    
    calculator = FormulaCalculator(sheet_data)
    
    # 测试各种公式
    test_formulas = [
        # 基本数学运算
        ("2+3*4", "基本数学运算"),
        ("100/5", "除法运算"),
        ("10^2", "幂运算"),
        
        # 单元格引用
        ("A2", "单元格引用"),
        ("B3+C3", "单元格相加"),
        
        # 函数调用
        ("SUM(1,2,3,4,5)", "SUM函数"),
        ("AVERAGE(10,20,30)", "AVERAGE函数"),
        ("MAX(5,15,8,20,3)", "MAX函数"),
        ("MIN(5,15,8,20,3)", "MIN函数"),
        ("COUNT(1,2,3)", "COUNT函数"),
        
        # 数学函数
        ("ABS(-25)", "ABS函数"),
        ("SQRT(16)", "SQRT函数"),
        ("ROUND(3.14159,2)", "ROUND函数"),
        ("POWER(2,8)", "POWER函数"),
        
        # 逻辑函数
        ('IF(10>5,"大","小")', "IF函数"),
        ("AND(1,1,0)", "AND函数"),
        ("OR(0,0,1)", "OR函数"),
        ("NOT(0)", "NOT函数"),
        
        # 文本函数
        ('LEN("Hello")', "LEN函数"),
        ('LEFT("Hello",3)', "LEFT函数"),
        ('RIGHT("World",3)', "RIGHT函数"),
        ('UPPER("hello")', "UPPER函数"),
        ('LOWER("WORLD")', "LOWER函数"),
        ('CONCATENATE("A","B","C")', "CONCATENATE函数"),
        
        # 错误测试
        ("10/0", "除零错误"),
        ("SQRT(-1)", "负数开方错误"),
        ("UNKNOWN(1,2)", "未知函数错误"),
    ]
    
    print(f"{'公式':<25} {'结果':<15} {'类型':<15} {'描述'}")
    print("-" * 80)
    
    for formula, description in test_formulas:
        try:
            result = calculator.calculate_formula(formula)
            
            # 格式化结果显示
            if result.error:
                result_str = f"❌ {result.error.value}"
            else:
                value = result.calculated_value
                if isinstance(value, float):
                    if value.is_integer():
                        result_str = str(int(value))
                    else:
                        result_str = f"{value:.6g}"
                else:
                    result_str = str(value)
            
            type_str = result.formula_type.value
            
            print(f"{formula:<25} {result_str:<15} {type_str:<15} {description}")
            
        except Exception as e:
            print(f"{formula:<25} {'错误':<15} {'error':<15} {str(e)}")


def demo_formula_processor():
    """演示公式处理器功能"""
    print("\n⚙️ === 公式处理器演示 ===")
    
    # 创建包含公式的测试数据
    sheet_data = {
        'data': [
            ['项目', '单价', '数量', '总价'],
            ['产品A', '100', '5', '=B2*C2'],
            ['产品B', '200', '3', '=B3*C3'],
            ['产品C', '150', '4', '=B4*C4'],
            ['合计', '', '', '=SUM(D2:D4)'],
            ['平均', '', '', '=AVERAGE(D2:D4)'],
            ['最大', '', '', '=MAX(D2:D4)'],
            ['计数', '', '', '=COUNT(D2:D4)']
        ],
        'styles': [
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}],
            [{}, {}, {}, {}]
        ]
    }
    
    # 创建处理器
    config = Config()
    processor = FormulaProcessor(config)
    
    # 处理公式
    enhanced_data = processor.process_sheet_formulas(sheet_data)
    
    # 显示结果
    print("发现的公式:")
    formulas = enhanced_data.get('formulas', {})
    
    for formula_key, formula_info in formulas.items():
        row, col = map(int, formula_key.split('_'))
        cell_ref = f"{chr(65+col)}{row+1}"
        
        print(f"  📍 {cell_ref}: {formula_info.original_formula}")
        print(f"     类型: {formula_info.formula_type.value}")
        print(f"     描述: {formula_info.description}")
        
        if formula_info.error:
            print(f"     ❌ 错误: {formula_info.error.value}")
        else:
            print(f"     ✅ 结果: {formula_info.calculated_value}")
        
        if formula_info.dependencies:
            print(f"     🔗 依赖: {', '.join(formula_info.dependencies)}")
        print()
    
    # 显示统计信息
    stats = processor.get_formula_statistics()
    print("📊 统计信息:")
    print(f"  总公式数: {stats['total_formulas']}")
    print(f"  成功计算: {stats['calculated_formulas']}")
    print(f"  错误公式: {stats['error_formulas']}")
    
    if stats['function_usage']:
        print("  函数使用:")
        for func, count in stats['function_usage'].items():
            print(f"    {func}: {count}次")
    
    if stats['error_types']:
        print("  错误类型:")
        for error, count in stats['error_types'].items():
            print(f"    {error}: {count}次")


def demo_complete_workflow():
    """演示完整的公式处理工作流"""
    print("\n🚀 === 完整工作流演示 ===")
    
    # 创建演示Excel文件
    excel_file = create_formula_demo_excel()
    
    # 解析Excel文件
    print("\n📊 解析Excel文件...")
    config = Config()
    config.ENABLE_FORMULA_PROCESSING = True
    
    parser = SheetParser(excel_file, config)
    sheets = parser.parse()
    
    if not sheets:
        print("❌ 解析失败")
        return
    
    sheet_data = sheets[0]
    print(f"✅ 解析完成，发现 {len(sheet_data.get('formulas', {}))} 个公式")
    
    # 创建HTML转换器
    print("\n🌐 转换为HTML...")
    converter = HTMLConverter(
        sheet_data,
        config=config,
        theme='default',
        use_css_classes=True
    )
    
    # 配置样式选项
    style_options = {
        'use_css_classes': True,
        'semantic_names': True,
        'min_usage_threshold': 1,
        'apply_conditional': False
    }
    
    # 生成HTML
    html_content = converter.to_html(style_options=style_options)
    
    # 保存HTML文件
    html_file = "demo_output/formula_demo.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✅ HTML文件已生成: {html_file}")
    
    # 显示公式统计
    formulas = sheet_data.get('formulas', {})
    if formulas:
        print(f"\n📈 公式分析:")
        
        # 按类型统计
        type_counts = {}
        error_counts = {}
        
        for formula_info in formulas.values():
            formula_type = formula_info.formula_type.value
            type_counts[formula_type] = type_counts.get(formula_type, 0) + 1
            
            if formula_info.error:
                error_name = formula_info.error.value
                error_counts[error_name] = error_counts.get(error_name, 0) + 1
        
        print("  公式类型分布:")
        for formula_type, count in type_counts.items():
            print(f"    {formula_type}: {count}个")
        
        if error_counts:
            print("  错误分布:")
            for error_type, count in error_counts.items():
                print(f"    {error_type}: {count}个")
        
        # 显示一些示例公式
        print("  示例公式:")
        count = 0
        for formula_key, formula_info in formulas.items():
            if count >= 5:  # 只显示前5个
                break
            
            row, col = map(int, formula_key.split('_'))
            cell_ref = f"{chr(65+col)}{row+1}"
            
            print(f"    {cell_ref}: {formula_info.original_formula}")
            if formula_info.error:
                print(f"         ❌ {formula_info.error.value}")
            else:
                print(f"         ✅ {formula_info.calculated_value}")
            
            count += 1
    
    return html_file


def main():
    """主演示函数"""
    print("🧮 公式处理功能演示")
    print("=" * 50)
    
    try:
        # 1. 公式计算器演示
        demo_formula_calculator()
        
        # 2. 公式处理器演示
        demo_formula_processor()
        
        # 3. 完整工作流演示
        html_file = demo_complete_workflow()
        
        print(f"\n✨ 演示完成！")
        print(f"📁 生成的文件:")
        print(f"  📊 Excel: demo_output/formula_demo.xlsx")
        print(f"  🌐 HTML: {html_file}")
        print(f"🌐 您可以在浏览器中打开HTML文件查看公式处理效果")
        print(f"💡 公式单元格有特殊的样式和悬停提示")
        
        return 0
        
    except Exception as e:
        print(f"\n❌ 演示过程中出现错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main()) 