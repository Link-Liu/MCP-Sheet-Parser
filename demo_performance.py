#!/usr/bin/env python3
"""
性能优化功能演示脚本
展示大文件处理、进度跟踪、内存监控等功能
"""

import os
import sys
import time
import openpyxl
from pathlib import Path

# 添加项目路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from mcp_sheet_parser.parser import SheetParser
from mcp_sheet_parser.html_converter import HTMLConverter
from mcp_sheet_parser.config import Config
from mcp_sheet_parser.performance import (
    PerformanceOptimizer, MemoryMonitor, create_progress_callback,
    benchmark_performance
)


def create_large_test_file(filename="large_test_file.xlsx", rows=5000, cols=10):
    """创建一个大的测试Excel文件"""
    print(f"🔨 创建测试文件: {filename} ({rows}行 × {cols}列)")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "性能测试数据"
    
    # 添加表头
    headers = [f"列{i+1}" for i in range(cols)]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 添加数据行
    for row in range(2, rows + 2):
        for col in range(1, cols + 1):
            # 生成一些测试数据
            if col == 1:
                ws.cell(row=row, column=col, value=f"数据行{row-1}")
            elif col == 2:
                ws.cell(row=row, column=col, value=row * 100)
            elif col == 3:
                ws.cell(row=row, column=col, value=f"=B{row}*2")  # 公式
            else:
                ws.cell(row=row, column=col, value=f"值{row}_{col}")
        
        # 每1000行显示进度
        if row % 1000 == 0:
            print(f"  创建进度: {row}/{rows+1} 行")
    
    # 添加一些样式和注释
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, color="FF0000")
    ws.cell(row=2, column=1).comment = openpyxl.comments.Comment("这是一个注释", "测试")
    
    # 添加超链接
    ws.cell(row=3, column=1).hyperlink = "https://example.com"
    ws.cell(row=3, column=1).value = "点击这里"
    
    # 保存文件
    wb.save(filename)
    wb.close()
    
    file_size = os.path.getsize(filename) / (1024 * 1024)
    print(f"✅ 文件创建完成: {file_size:.2f}MB")
    return filename


def demo_system_info():
    """演示系统信息获取"""
    print("\n🖥️  系统信息:")
    optimizer = PerformanceOptimizer()
    info = optimizer.get_system_info()
    
    print(f"  CPU核心数: {info['cpu_count']}")
    print(f"  CPU使用率: {info['cpu_percent']:.1f}%")
    print(f"  总内存: {info['memory_total_gb']:.1f}GB")
    print(f"  可用内存: {info['memory_available_gb']:.1f}GB")
    print(f"  内存使用率: {info['memory_percent']:.1f}%")


def demo_memory_monitor():
    """演示内存监控功能"""
    print("\n🧠 内存监控演示:")
    
    monitor = MemoryMonitor(max_memory_mb=1024)
    print(f"  {monitor.get_memory_stats()}")
    
    # 分配一些内存来演示监控
    print("  分配测试内存...")
    test_data = [i for i in range(1000000)]  # 分配一些内存
    
    print(f"  {monitor.get_memory_stats()}")
    
    if monitor.should_trigger_gc():
        print("  触发垃圾回收...")
        freed = monitor.trigger_gc()
        print(f"  释放内存: {freed:.2f}MB")
    
    del test_data  # 释放内存


def demo_progress_tracking():
    """演示进度跟踪功能"""
    print("\n📊 进度跟踪演示:")
    
    from mcp_sheet_parser.performance import ProgressTracker
    
    tracker = ProgressTracker(100, "演示处理")
    
    # 添加进度回调
    def progress_callback(info):
        print(f"\r  {info['description']}: {info['percentage']:.1f}% "
              f"({info['processed']}/{info['total']}) "
              f"- {info['speed']} - 剩余: {info['eta']}", end='')
    
    tracker.add_callback(progress_callback)
    
    # 模拟处理过程
    for i in range(100):
        time.sleep(0.01)  # 模拟工作
        tracker.update(1)
    
    print("\n  ✅ 进度跟踪完成")


def demo_performance_parsing(test_file):
    """演示性能优化解析"""
    print(f"\n🚀 性能优化解析演示: {test_file}")
    
    # 创建配置
    config = Config()
    config.ENABLE_PERFORMANCE_MODE = True
    config.CHUNK_SIZE = 1000
    config.MAX_MEMORY_MB = 2048
    config.ENABLE_PROGRESS_TRACKING = True
    
    print(f"  配置: 块大小={config.CHUNK_SIZE}, 最大内存={config.MAX_MEMORY_MB}MB")
    
    # 创建进度回调
    progress_callback = create_progress_callback(verbose=True)
    
    # 性能基准测试
    def parse_with_performance():
        parser = SheetParser(test_file, config, progress_callback)
        return parser.parse()
    
    print("  开始性能基准测试...")
    result = benchmark_performance(parse_with_performance)
    
    if result['success']:
        sheets = result['result']
        print(f"\n  ✅ 解析完成:")
        print(f"    执行时间: {result['execution_time']:.2f}秒")
        print(f"    内存变化: {result['memory_delta_mb']:.2f}MB")
        print(f"    峰值内存: {result['peak_memory_mb']:.2f}MB")
        print(f"    工作表数量: {len(sheets)}")
        
        for i, sheet in enumerate(sheets):
            print(f"    工作表{i}: {sheet['sheet_name']} "
                  f"({sheet['rows']}行 × {sheet['cols']}列)")
        
        return sheets
    else:
        print(f"  ❌ 解析失败: {result['error']}")
        return None


def demo_html_conversion(sheets, output_file="performance_demo.html"):
    """演示HTML转换"""
    if not sheets:
        return
    
    print(f"\n🎨 HTML转换演示: {output_file}")
    
    # 使用第一个工作表
    sheet = sheets[0]
    
    # 创建HTML转换器
    converter = HTMLConverter(sheet, theme='default')
    
    # 转换并保存
    converter.export_to_file(output_file)
    
    file_size = os.path.getsize(output_file) / 1024
    print(f"  ✅ HTML文件已生成: {file_size:.1f}KB")
    
    # 显示一些统计信息
    if sheet.get('comments'):
        print(f"  📝 包含注释: {len(sheet['comments'])}个")
    
    if sheet.get('hyperlinks'):
        print(f"  🔗 包含超链接: {len(sheet['hyperlinks'])}个")
    
    if sheet.get('merged_cells'):
        print(f"  🔗 合并单元格: {len(sheet['merged_cells'])}个")


def demo_comparison(test_file):
    """演示性能对比"""
    print(f"\n⚖️  性能对比演示:")
    
    # 标准模式
    print("  测试标准解析模式...")
    config_standard = Config()
    config_standard.ENABLE_PERFORMANCE_MODE = False
    
    def parse_standard():
        parser = SheetParser(test_file, config_standard)
        return parser.parse()
    
    result_standard = benchmark_performance(parse_standard)
    
    # 性能优化模式
    print("  测试性能优化模式...")
    config_performance = Config()
    config_performance.ENABLE_PERFORMANCE_MODE = True
    config_performance.CHUNK_SIZE = 500
    
    def parse_performance():
        parser = SheetParser(test_file, config_performance)
        return parser.parse()
    
    result_performance = benchmark_performance(parse_performance)
    
    # 对比结果
    print("\n  📈 性能对比结果:")
    print(f"    标准模式:")
    print(f"      执行时间: {result_standard['execution_time']:.2f}秒")
    print(f"      内存使用: {result_standard['memory_delta_mb']:.2f}MB")
    
    print(f"    优化模式:")
    print(f"      执行时间: {result_performance['execution_time']:.2f}秒")
    print(f"      内存使用: {result_performance['memory_delta_mb']:.2f}MB")
    
    # 计算改进比例
    time_improvement = ((result_standard['execution_time'] - result_performance['execution_time']) 
                       / result_standard['execution_time'] * 100)
    memory_improvement = ((result_standard['memory_delta_mb'] - result_performance['memory_delta_mb']) 
                         / result_standard['memory_delta_mb'] * 100) if result_standard['memory_delta_mb'] > 0 else 0
    
    print(f"\n  🎯 性能改进:")
    print(f"    时间优化: {time_improvement:+.1f}%")
    print(f"    内存优化: {memory_improvement:+.1f}%")


def main():
    """主演示函数"""
    print("🎯 MCP-Sheet-Parser 性能优化功能演示")
    print("=" * 50)
    
    try:
        # 1. 系统信息
        demo_system_info()
        
        # 2. 内存监控
        demo_memory_monitor()
        
        # 3. 进度跟踪
        demo_progress_tracking()
        
        # 4. 创建测试文件
        test_file = create_large_test_file(rows=3000, cols=8)  # 适中的大小
        
        # 5. 性能优化解析
        sheets = demo_performance_parsing(test_file)
        
        # 6. HTML转换
        demo_html_conversion(sheets)
        
        # 7. 性能对比（可选，较耗时）
        if input("\n❓ 是否运行性能对比测试？(y/N): ").lower() == 'y':
            demo_comparison(test_file)
        
        print(f"\n🎉 演示完成！")
        print(f"📁 生成的文件:")
        print(f"  - 测试文件: {test_file}")
        print(f"  - HTML输出: performance_demo.html")
        
        # 清理询问
        if input("\n🗑️  是否删除测试文件？(y/N): ").lower() == 'y':
            os.remove(test_file)
            print("✅ 测试文件已删除")
        
    except KeyboardInterrupt:
        print("\n⏹️  用户中断演示")
    except Exception as e:
        print(f"\n❌ 演示过程中出现错误: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main() 