#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
图表转换功能演示程序
展示SVG图表生成和HTML集成功能
"""

import os
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from mcp_sheet_parser.parser import SheetParser
from mcp_sheet_parser.html_converter import HTMLConverter
from mcp_sheet_parser.chart_converter import (
    ChartConverter, ChartData, ChartType
)
from mcp_sheet_parser.config import Config


def create_chart_demo_excel():
    """创建包含数据的演示Excel文件用于图表生成"""
    print("📊 创建图表演示Excel文件...")
    
    wb = openpyxl.Workbook()
    
    # 第一个工作表：销售数据
    ws1 = wb.active
    ws1.title = "季度销售数据"
    
    # 设置表头
    headers = ['季度', '产品A', '产品B', '产品C', '总计']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 添加销售数据
    sales_data = [
        ['Q1', 120, 100, 80, '=B2+C2+D2'],
        ['Q2', 150, 130, 110, '=B3+C3+D3'],
        ['Q3', 180, 160, 140, '=B4+C4+D4'],
        ['Q4', 200, 170, 150, '=B5+C5+D5']
    ]
    
    for row_idx, row_data in enumerate(sales_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # 第二个工作表：市场份额数据
    ws2 = wb.create_sheet("市场份额数据")
    
    # 设置表头
    share_headers = ['区域', '份额(%)']
    for col, header in enumerate(share_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 添加市场份额数据
    share_data = [
        ['华东', 35],
        ['华南', 25],
        ['华北', 20],
        ['华中', 15],
        ['其他', 5]
    ]
    
    for row_idx, row_data in enumerate(share_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=value)
    
    # 第三个工作表：月度趋势数据
    ws3 = wb.create_sheet("月度趋势数据")
    
    # 设置表头
    trend_headers = ['月份', '收入(万元)', '利润(万元)', '增长率(%)']
    for col, header in enumerate(trend_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # 添加月度趋势数据
    trend_data = [
        ['1月', 100, 20, 0],
        ['2月', 120, 25, 20],
        ['3月', 110, 22, -8.3],
        ['4月', 140, 30, 27.3],
        ['5月', 160, 35, 14.3],
        ['6月', 180, 40, 12.5]
    ]
    
    for row_idx, row_data in enumerate(trend_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    output_path = "demo_output/chart_demo.xlsx"
    os.makedirs("demo_output", exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Excel文件已创建: {output_path}")
    return output_path


def demo_chart_generators():
    """演示各种图表生成器"""
    print("\n📊 === 图表生成器演示 ===")
    
    converter = ChartConverter()
    
    # 柱状图演示
    print("\n📊 生成柱状图...")
    column_chart = ChartData(
        chart_type=ChartType.COLUMN,
        title="季度销售业绩对比",
        categories=["Q1", "Q2", "Q3", "Q4"],
        data_series=[
            {"name": "产品A", "values": [120, 150, 180, 200]},
            {"name": "产品B", "values": [100, 130, 160, 170]},
            {"name": "产品C", "values": [80, 110, 140, 150]}
        ],
        x_axis_title="季度",
        y_axis_title="销售额(万元)",
        width=700,
        height=400
    )
    
    column_svg = converter.generate_svg(column_chart)
    
    # 保存柱状图
    with open("demo_output/column_chart.svg", 'w', encoding='utf-8') as f:
        f.write(column_svg)
    print("✅ 柱状图已保存: demo_output/column_chart.svg")
    
    # 饼图演示
    print("\n🥧 生成饼图...")
    pie_chart = ChartData(
        chart_type=ChartType.PIE,
        title="市场份额分布",
        categories=["华东", "华南", "华北", "华中", "其他"],
        data_series=[
            {"name": "市场份额", "values": [35, 25, 20, 15, 5]}
        ],
        width=500,
        height=400
    )
    
    pie_svg = converter.generate_svg(pie_chart)
    
    # 保存饼图
    with open("demo_output/pie_chart.svg", 'w', encoding='utf-8') as f:
        f.write(pie_svg)
    print("✅ 饼图已保存: demo_output/pie_chart.svg")
    
    # 折线图演示
    print("\n📈 生成折线图...")
    line_chart = ChartData(
        chart_type=ChartType.LINE,
        title="月度增长趋势",
        categories=["1月", "2月", "3月", "4月", "5月", "6月"],
        data_series=[
            {"name": "收入", "values": [100, 120, 110, 140, 160, 180]},
            {"name": "利润", "values": [20, 25, 22, 30, 35, 40]}
        ],
        x_axis_title="月份",
        y_axis_title="金额(万元)",
        width=700,
        height=400
    )
    
    line_svg = converter.generate_svg(line_chart)
    
    # 保存折线图
    with open("demo_output/line_chart.svg", 'w', encoding='utf-8') as f:
        f.write(line_svg)
    print("✅ 折线图已保存: demo_output/line_chart.svg")
    
    return [column_chart, pie_chart, line_chart]


def demo_integrated_html_with_charts():
    """演示集成图表的完整HTML工作流"""
    print("\n🚀 === 完整HTML+图表集成演示 ===")
    
    # 创建演示Excel文件
    excel_file = create_chart_demo_excel()
    
    # 解析Excel文件
    print("\n📊 解析Excel文件...")
    config = Config()
    config.ENABLE_CHART_CONVERSION = True
    
    parser = SheetParser(excel_file, config)
    sheets = parser.parse()
    
    if not sheets:
        print("❌ 解析失败")
        return
    
    # 处理每个工作表
    html_files = []
    for i, sheet_data in enumerate(sheets):
        sheet_name = sheet_data['sheet_name']
        print(f"\n🌐 处理工作表: {sheet_name}")
        
        # 显示图表信息
        if 'charts' in sheet_data:
            charts = sheet_data['charts']
            print(f"✅ 发现 {len(charts)} 个图表")
            for j, chart in enumerate(charts):
                print(f"  图表 {j+1}: {chart['title']} ({chart['type']})")
        
        # 创建HTML转换器
        converter = HTMLConverter(
            sheet_data,
            config=config,
            theme='default',
            use_css_classes=True
        )
        
        # 生成HTML
        html_content = converter.to_html()
        
        # 保存HTML文件
        html_file = f"demo_output/chart_demo_sheet{i+1}.html"
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        html_files.append(html_file)
        print(f"✅ HTML文件已生成: {html_file}")
    
    return html_files


def create_chart_gallery():
    """创建图表展示画廊"""
    print("\n🎨 === 创建图表展示画廊 ===")
    
    converter = ChartConverter()
    demo_charts = converter.create_demo_charts()
    
    # 生成所有图表的SVG
    chart_svgs = []
    for i, chart_data in enumerate(demo_charts):
        svg_content = converter.generate_svg(chart_data)
        chart_svgs.append({
            'title': chart_data.title,
            'type': chart_data.chart_type.value,
            'svg': svg_content
        })
    
    # 创建HTML画廊
    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>📊 MCP-Sheet-Parser 图表展示画廊</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: #333;
        }}
        
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.2);
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #4472C4 0%, #2E4BC6 100%);
            color: white;
            padding: 30px;
            text-align: center;
        }}
        
        .header h1 {{
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }}
        
        .header p {{
            margin: 10px 0 0;
            font-size: 1.2em;
            opacity: 0.9;
        }}
        
        .gallery {{
            padding: 40px;
        }}
        
        .chart-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 30px;
        }}
        
        .chart-card {{
            background: #f8f9fa;
            border-radius: 12px;
            padding: 25px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
        }}
        
        .chart-card:hover {{
            transform: translateY(-5px);
            box-shadow: 0 8px 25px rgba(0,0,0,0.15);
        }}
        
        .chart-title {{
            font-size: 1.4em;
            font-weight: 600;
            color: #495057;
            margin-bottom: 15px;
            text-align: center;
        }}
        
        .chart-content {{
            text-align: center;
            overflow-x: auto;
        }}
        
        .chart-content svg {{
            max-width: 100%;
            height: auto;
            border-radius: 8px;
        }}
        
        .chart-meta {{
            margin-top: 15px;
            text-align: center;
            color: #6c757d;
            font-size: 0.9em;
        }}
        
        .footer {{
            background: #f8f9fa;
            padding: 20px;
            text-align: center;
            color: #6c757d;
            border-top: 1px solid #dee2e6;
        }}
        
        @media (max-width: 768px) {{
            .chart-grid {{
                grid-template-columns: 1fr;
            }}
            
            .header h1 {{
                font-size: 2em;
            }}
            
            .header p {{
                font-size: 1em;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 图表展示画廊</h1>
            <p>MCP-Sheet-Parser v2.3.0 - 专业Excel图表转换演示</p>
        </div>
        
        <div class="gallery">
            <div class="chart-grid">
"""
    
    for chart in chart_svgs:
        html_content += f"""
                <div class="chart-card">
                    <div class="chart-title">{chart['title']}</div>
                    <div class="chart-content">
                        {chart['svg']}
                    </div>
                    <div class="chart-meta">
                        类型: {chart['type']} | 格式: SVG矢量图
                    </div>
                </div>
"""
    
    html_content += """
            </div>
        </div>
        
        <div class="footer">
            <p>🚀 Powered by MCP-Sheet-Parser | 支持柱状图、饼图、折线图等多种图表类型</p>
            <p>💡 图表采用SVG格式，支持缩放和响应式显示</p>
        </div>
    </div>
    
    <script>
        // 添加图表交互效果
        document.addEventListener('DOMContentLoaded', function() {
            const charts = document.querySelectorAll('.chart-content svg');
            
            charts.forEach(chart => {
                // 添加悬停效果
                chart.addEventListener('mouseenter', function() {
                    this.style.opacity = '0.9';
                    this.style.transform = 'scale(1.02)';
                    this.style.transition = 'all 0.3s ease';
                });
                
                chart.addEventListener('mouseleave', function() {
                    this.style.opacity = '1';
                    this.style.transform = 'scale(1)';
                });
            });
        });
    </script>
</body>
</html>"""
    
    # 保存画廊文件
    gallery_file = "demo_output/chart_gallery.html"
    with open(gallery_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✅ 图表画廊已创建: {gallery_file}")
    return gallery_file


def main():
    """主演示函数"""
    print("📊 图表转换功能演示")
    print("=" * 50)
    
    try:
        # 确保输出目录存在
        os.makedirs("demo_output", exist_ok=True)
        
        # 1. 图表生成器演示
        demo_charts = demo_chart_generators()
        
        # 2. 集成HTML工作流演示
        html_files = demo_integrated_html_with_charts()
        
        # 3. 创建图表展示画廊
        gallery_file = create_chart_gallery()
        
        print(f"\n✨ 图表转换演示完成！")
        print(f"📁 生成的文件:")
        print(f"  📊 Excel数据: demo_output/chart_demo.xlsx")
        print(f"  📈 独立图表: demo_output/column_chart.svg, pie_chart.svg, line_chart.svg")
        print(f"  🌐 集成HTML: {', '.join(html_files)}")
        print(f"  🎨 图表画廊: {gallery_file}")
        
        print(f"\n🌟 特色功能:")
        print(f"  ✅ 支持柱状图、饼图、折线图等多种类型")
        print(f"  ✅ SVG矢量图格式，支持缩放不失真")
        print(f"  ✅ 响应式设计，适配不同屏幕尺寸")
        print(f"  ✅ 专业的配色方案和视觉效果")
        print(f"  ✅ 完美集成到HTML表格中")
        
        print(f"\n💡 使用建议:")
        print(f"  🌐 在浏览器中打开HTML文件查看效果")
        print(f"  📱 尝试在不同设备上查看响应式效果")
        print(f"  🎨 查看图表画廊了解所有支持的图表类型")
        
        return 0
        
    except Exception as e:
        print(f"\n❌ 演示过程中出现错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    exit(main()) 