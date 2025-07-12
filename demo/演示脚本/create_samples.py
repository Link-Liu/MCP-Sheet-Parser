#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建示例文件脚本
用于生成各种格式的示例文件，展示MCP-Sheet-Parser的功能
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import csv

def create_directory_structure():
    """创建目录结构"""
    directories = [
        '示例文件/excel',
        '示例文件/csv', 
        '示例文件/wps',
        '示例文件/complex'
    ]
    
    for directory in directories:
        os.makedirs(directory, exist_ok=True)
        print(f"✅ 创建目录: {directory}")

def create_basic_data():
    """创建基础数据"""
    return {
        '姓名': ['张三', '李四', '王五', '赵六', '钱七'],
        '年龄': [25, 30, 35, 28, 32],
        '部门': ['技术部', '市场部', '人事部', '财务部', '技术部'],
        '薪资': [8000, 12000, 9000, 10000, 15000],
        '入职日期': ['2020-01-15', '2019-03-20', '2018-07-10', '2021-02-28', '2017-11-05']
    }

def create_complex_data():
    """创建复杂数据（包含公式、样式等）"""
    return {
        '产品': ['产品A', '产品B', '产品C', '产品D', '产品E'],
        '单价': [100, 200, 150, 300, 250],
        '数量': [10, 5, 8, 3, 12],
        '折扣': [0.1, 0.2, 0.15, 0.05, 0.25],
        '小计': ['=B2*C2*(1-D2)', '=B3*C3*(1-D3)', '=B4*C4*(1-D4)', '=B5*C5*(1-D5)', '=B6*C6*(1-D6)'],
        '备注': ['热销产品', '新品上市', '库存充足', '限量供应', '促销中']
    }

def create_csv_samples():
    """创建CSV示例文件"""
    print("\n📄 创建CSV示例文件...")
    
    # 基础CSV
    df_basic = pd.DataFrame(create_basic_data())
    df_basic.to_csv('示例文件/csv/basic_sample.csv', index=False, encoding='utf-8-sig')
    print("✅ 创建: basic_sample.csv")
    
    # 复杂CSV（包含特殊字符）
    df_complex = pd.DataFrame(create_complex_data())
    df_complex.to_csv('示例文件/csv/complex_sample.csv', index=False, encoding='utf-8-sig')
    print("✅ 创建: complex_sample.csv")
    
    # 多语言CSV
    multilingual_data = {
        'Name/姓名': ['John/约翰', 'Mary/玛丽', 'Tom/汤姆'],
        'Age/年龄': [25, 30, 35],
        'Department/部门': ['IT/技术', 'Sales/销售', 'HR/人事']
    }
    df_multilingual = pd.DataFrame(multilingual_data)
    df_multilingual.to_csv('示例文件/csv/multilingual_sample.csv', index=False, encoding='utf-8-sig')
    print("✅ 创建: multilingual_sample.csv")

def create_excel_samples():
    """创建Excel示例文件"""
    print("\n📊 创建Excel示例文件...")
    
    # 基础Excel文件
    wb_basic = openpyxl.Workbook()
    ws_basic = wb_basic.active
    ws_basic.title = "员工信息"
    
    # 添加数据
    data = create_basic_data()
    headers = list(data.keys())
    ws_basic.append(headers)
    
    for i in range(len(data[headers[0]])):
        row = [data[col][i] for col in headers]
        ws_basic.append(row)
    
    # 添加样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_basic[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    wb_basic.save('示例文件/excel/basic_sample.xlsx')
    print("✅ 创建: basic_sample.xlsx")
    
    # 复杂Excel文件（包含公式、样式、批注等）
    wb_complex = openpyxl.Workbook()
    ws_complex = wb_complex.active
    ws_complex.title = "销售数据"
    
    # 添加数据
    complex_data = create_complex_data()
    headers = list(complex_data.keys())
    ws_complex.append(headers)
    
    for i in range(len(complex_data[headers[0]])):
        row = [complex_data[col][i] for col in headers]
        ws_complex.append(row)
    
    # 添加公式
    for i in range(2, 7):
        ws_complex[f'E{i}'] = f'=B{i}*C{i}*(1-D{i})'
    
    # 添加样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
    
    for cell in ws_complex[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # 添加批注
    ws_complex['F2'].comment = openpyxl.comments.Comment("这是热销产品", "系统")
    ws_complex['F3'].comment = openpyxl.comments.Comment("新品上市，需要推广", "系统")
    
    # 添加超链接
    ws_complex['A1'].hyperlink = "https://example.com"
    ws_complex['A1'].value = "产品信息 (点击查看详情)"
    
    wb_complex.save('示例文件/excel/complex_sample.xlsx')
    print("✅ 创建: complex_sample.xlsx")
    
    # Excel模板文件
    wb_template = openpyxl.Workbook()
    ws_template = wb_template.active
    ws_template.title = "月度报告模板"
    
    # 添加模板结构
    template_data = [
        ["月度销售报告", "", "", "", ""],
        ["", "", "", "", ""],
        ["部门", "目标", "实际", "完成率", "备注"],
        ["技术部", 100000, "", "=C5/B5", ""],
        ["市场部", 150000, "", "=C6/B6", ""],
        ["销售部", 200000, "", "=C7/B7", ""],
        ["", "", "", "", ""],
        ["总计", "=SUM(B5:B7)", "=SUM(C5:C7)", "=C9/B9", ""]
    ]
    
    for row in template_data:
        ws_template.append(row)
    
    # 添加样式
    title_font = Font(bold=True, size=16)
    ws_template['A1'].font = title_font
    
    wb_template.save('示例文件/excel/template_sample.xltx')
    print("✅ 创建: template_sample.xltx")

def create_wps_samples():
    """创建WPS示例文件（使用Excel格式，但扩展名改为WPS格式）"""
    print("\n📝 创建WPS示例文件...")
    
    # 由于WPS格式的特殊性，这里创建Excel格式但使用WPS扩展名
    # 实际项目中需要专门的WPS处理库
    
    # WPS表格文件
    wb_wps = openpyxl.Workbook()
    ws_wps = wb_wps.active
    ws_wps.title = "WPS表格示例"
    
    data = create_basic_data()
    headers = list(data.keys())
    ws_wps.append(headers)
    
    for i in range(len(data[headers[0]])):
        row = [data[col][i] for col in headers]
        ws_wps.append(row)
    
    wb_wps.save('示例文件/wps/wps_sample.et')
    print("✅ 创建: wps_sample.et")
    
    # WPS模板文件
    wb_wps_template = openpyxl.Workbook()
    ws_wps_template = wb_wps_template.active
    ws_wps_template.title = "WPS模板"
    
    template_data = [
        ["WPS表格模板", "", "", ""],
        ["", "", "", ""],
        ["项目", "预算", "实际", "差异"],
        ["项目A", 50000, "", "=C5-B5"],
        ["项目B", 30000, "", "=C6-B6"],
        ["项目C", 20000, "", "=C7-B7"]
    ]
    
    for row in template_data:
        ws_wps_template.append(row)
    
    wb_wps_template.save('示例文件/wps/wps_template.ett')
    print("✅ 创建: wps_template.ett")

def create_complex_samples():
    """创建复杂示例文件"""
    print("\n🔧 创建复杂示例文件...")
    
    # 包含多个工作表的Excel文件
    wb_multi = openpyxl.Workbook()
    
    # 第一个工作表：员工信息
    ws_employees = wb_multi.active
    ws_employees.title = "员工信息"
    
    employee_data = create_basic_data()
    headers = list(employee_data.keys())
    ws_employees.append(headers)
    
    for i in range(len(employee_data[headers[0]])):
        row = [employee_data[col][i] for col in headers]
        ws_employees.append(row)
    
    # 第二个工作表：销售数据
    ws_sales = wb_multi.create_sheet("销售数据")
    
    sales_data = create_complex_data()
    headers = list(sales_data.keys())
    ws_sales.append(headers)
    
    for i in range(len(sales_data[headers[0]])):
        row = [sales_data[col][i] for col in headers]
        ws_sales.append(row)
    
    # 第三个工作表：统计图表
    ws_stats = wb_multi.create_sheet("统计信息")
    
    stats_data = [
        ["统计项目", "数值", "百分比"],
        ["总员工数", 5, "100%"],
        ["技术部", 2, "40%"],
        ["市场部", 1, "20%"],
        ["人事部", 1, "20%"],
        ["财务部", 1, "20%"]
    ]
    
    for row in stats_data:
        ws_stats.append(row)
    
    wb_multi.save('示例文件/complex/multi_sheet_sample.xlsx')
    print("✅ 创建: multi_sheet_sample.xlsx")
    
    # 包含合并单元格的文件
    wb_merged = openpyxl.Workbook()
    ws_merged = wb_merged.active
    ws_merged.title = "合并单元格示例"
    
    # 添加数据
    merged_data = [
        ["部门", "姓名", "年龄", "薪资"],
        ["技术部", "张三", 25, 8000],
        ["", "李四", 32, 15000],
        ["市场部", "王五", 30, 12000],
        ["人事部", "赵六", 28, 10000],
        ["财务部", "钱七", 35, 9000]
    ]
    
    for row in merged_data:
        ws_merged.append(row)
    
    # 合并单元格
    ws_merged.merge_cells('A2:A3')  # 技术部
    ws_merged.merge_cells('A4:A4')  # 市场部
    ws_merged.merge_cells('A5:A5')  # 人事部
    ws_merged.merge_cells('A6:A6')  # 财务部
    
    wb_merged.save('示例文件/complex/merged_cells_sample.xlsx')
    print("✅ 创建: merged_cells_sample.xlsx")

def create_readme():
    """创建示例文件说明文档"""
    readme_content = """# 示例文件说明

本目录包含MCP-Sheet-Parser的各种示例文件，用于展示和测试不同格式的表格文件转换功能。

## 文件结构

### Excel格式示例
- `basic_sample.xlsx` - 基础Excel文件，包含员工信息
- `complex_sample.xlsx` - 复杂Excel文件，包含公式、样式、批注、超链接
- `template_sample.xltx` - Excel模板文件

### CSV格式示例
- `basic_sample.csv` - 基础CSV文件
- `complex_sample.csv` - 复杂CSV文件，包含特殊字符
- `multilingual_sample.csv` - 多语言CSV文件

### WPS格式示例
- `wps_sample.et` - WPS表格文件
- `wps_template.ett` - WPS模板文件

### 复杂示例
- `multi_sheet_sample.xlsx` - 多工作表Excel文件
- `merged_cells_sample.xlsx` - 包含合并单元格的Excel文件

## 使用说明

1. 这些示例文件可以用于测试MCP-Sheet-Parser的各种功能
2. 每个文件都包含不同的数据结构和格式特性
3. 可以通过命令行工具转换这些文件：
   ```bash
   python main.py 示例文件/excel/basic_sample.xlsx --html output.html
   ```

## 文件特性

### 基础功能
- 文本、数字、日期数据类型
- 基本的表格结构
- 简单的样式信息

### 高级功能
- 公式计算
- 单元格合并
- 批注信息
- 超链接
- 多工作表
- 复杂样式

### 特殊功能
- 多语言支持
- 模板功能
- 特殊字符处理
- 大数据量处理

这些示例文件涵盖了MCP-Sheet-Parser支持的所有主要功能，是测试和演示的理想选择。
"""
    
    with open('示例文件/README.md', 'w', encoding='utf-8') as f:
        f.write(readme_content)
    
    print("✅ 创建: README.md")

def main():
    """主函数"""
    print("🚀 开始创建示例文件...")
    print("=" * 50)
    
    try:
        # 创建目录结构
        create_directory_structure()
        
        # 创建各种格式的示例文件
        create_csv_samples()
        create_excel_samples()
        create_wps_samples()
        create_complex_samples()
        
        # 创建说明文档
        create_readme()
        
        print("\n" + "=" * 50)
        print("🎉 示例文件创建完成！")
        print("\n📁 创建的文件:")
        print("├── 示例文件/excel/")
        print("│   ├── basic_sample.xlsx")
        print("│   ├── complex_sample.xlsx")
        print("│   └── template_sample.xltx")
        print("├── 示例文件/csv/")
        print("│   ├── basic_sample.csv")
        print("│   ├── complex_sample.csv")
        print("│   └── multilingual_sample.csv")
        print("├── 示例文件/wps/")
        print("│   ├── wps_sample.et")
        print("│   └── wps_template.ett")
        print("├── 示例文件/complex/")
        print("│   ├── multi_sheet_sample.xlsx")
        print("│   └── merged_cells_sample.xlsx")
        print("└── 示例文件/README.md")
        
        print("\n💡 提示: 这些示例文件可以用于测试MCP-Sheet-Parser的各种功能")
        
    except Exception as e:
        print(f"❌ 创建示例文件时出错: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 